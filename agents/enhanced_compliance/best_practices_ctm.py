"""
PropelAI v2.12: Best Practices CTM Exporter

Per best practices, this creates THREE distinct compliance matrices:

A. "Section L Compliance Matrix"
   - Used internally to ensure your volume, headings, page limits, 
     and submission format follow the instructions exactly.

B. "Technical Requirements Compliance Matrix" 
   - Evaluator-facing (if submitted), proving you meet 100% of 
     mandatory requirements from C/PWS/SOW.

C. "Section M Alignment Matrix"
   - Shows evaluators exactly where you meet or exceed their 
     scoring criteria.

Each matrix uses the federal-level recommended 5-6 column model:
| Requirement ID | Requirement Text (Verbatim) | Source | Proposal Location | Compliance Response | Notes |
"""

import os
from typing import List, Dict, Any, Optional
from datetime import datetime

# openpyxl imports moved to __init__ to handle installation timing issues
OPENPYXL_AVAILABLE = False

from .section_aware_extractor import (
    StructuredRequirement,
    ExtractionResult,
    RequirementCategory,
    BindingLevel
)


class BestPracticesCTMExporter:
    """
    Export CTM following best practices from federal proposal experts.
    
    Creates a professional workbook with:
    1. Cover Sheet - Summary and navigation
    2. Section L Compliance - Submission instructions
    3. Technical Requirements - C/PWS/SOW requirements
    4. Section M Alignment - Evaluation factors
    5. All Requirements - Complete list for reference
    6. Cross-Reference Matrix - Links between L/M/C
    """
    
    # Color scheme (professional, not garish)
    COLORS = {
        'header': '1F4E79',          # Dark blue
        'mandatory': 'FCE4D6',       # Light orange tint
        'desirable': 'E2EFDA',       # Light green tint
        'l_section': 'D9E2F3',       # Light blue (instructions)
        'm_section': 'FFF2CC',       # Light yellow (evaluation)
        'c_section': 'E2EFDA',       # Light green (technical)
        'att_section': 'F2F2F2',     # Light gray (attachments)
        'response_col': 'FFFFFF',    # White (for team input)
        'alternate_row': 'F9F9F9',   # Very light gray
        'high_priority': 'F8CBAD',   # Orange for HIGH
        'medium_priority': 'FFE699', # Yellow for MEDIUM
        'low_priority': 'C6EFCE',    # Green for LOW
    }
    
    # Priority mapping from binding level
    PRIORITY_MAP = {
        BindingLevel.MANDATORY: "High",
        BindingLevel.HIGHLY_DESIRABLE: "Medium",
        BindingLevel.DESIRABLE: "Low",
        BindingLevel.INFORMATIONAL: "Low",
    }
    
    def __init__(self, include_response_columns: bool = True):
        """
        Args:
            include_response_columns: If True, include columns for team to fill in
        """
        self.include_response_columns = include_response_columns
        
        # Import openpyxl at runtime and make classes available globally in module
        try:
            global Workbook, Font, Fill, PatternFill, Alignment, Border, Side, DataValidation, OPENPYXL_AVAILABLE
            from openpyxl import Workbook
            from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            OPENPYXL_AVAILABLE = True
        except ImportError:
            raise ImportError("openpyxl is required for CTM export")
    
    def _get_priority(self, binding_level: BindingLevel) -> str:
        """Map binding level to priority string"""
        return self.PRIORITY_MAP.get(binding_level, "Medium")
    
    def _get_priority_color(self, priority: str) -> str:
        """Get color for priority level"""
        color_map = {
            "High": self.COLORS['high_priority'],
            "Medium": self.COLORS['medium_priority'],
            "Low": self.COLORS['low_priority'],
        }
        return color_map.get(priority, self.COLORS['medium_priority'])

    def _safe_cell_value(self, value: str) -> str:
        """
        Escape cell values that Excel would interpret as formulas.

        Excel treats cells starting with =, +, -, @, or tab as formulas.
        Prefix with single quote to force text interpretation.
        """
        if not value or not isinstance(value, str):
            return value

        # Characters that Excel interprets as formula starters
        formula_chars = ('=', '+', '-', '@', '\t')

        if value.strip().startswith(formula_chars):
            return "'" + value

        return value

    def export(self, result: ExtractionResult, output_path: str, 
               solicitation_number: str = "", title: str = "") -> str:
        """
        Export extraction result to Excel CTM.
        
        Args:
            result: ExtractionResult from SectionAwareExtractor
            output_path: Where to save the Excel file
            solicitation_number: RFP/Solicitation number for header
            title: RFP title for header
            
        Returns:
            Path to created file
        """
        wb = Workbook()
        
        # Use structure info if available
        if result.structure:
            solicitation_number = solicitation_number or result.structure.solicitation_number
            title = title or result.structure.title
        
        # Create sheets
        self._create_cover_sheet(wb, result, solicitation_number, title)
        self._create_section_l_matrix(wb, result.section_l_requirements)
        self._create_technical_matrix(wb, result.technical_requirements)
        self._create_section_m_matrix(wb, result.evaluation_requirements)
        self._create_all_requirements_sheet(wb, result.all_requirements)
        
        if result.structure:
            self._create_cross_reference_sheet(wb, result)
        
        # Remove default sheet if it exists
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Save
        wb.save(output_path)
        return output_path
    
    def _create_cover_sheet(self, wb: "Workbook", result: ExtractionResult, 
                            solicitation_number: str, title: str):
        """Create summary cover sheet"""
        ws = wb.create_sheet("Cover Sheet", 0)
        
        # Title
        ws['A1'] = "COMPLIANCE TRACEABILITY MATRIX"
        ws['A1'].font = Font(size=18, bold=True, color=self.COLORS['header'])
        
        ws['A3'] = f"Solicitation: {solicitation_number}"
        ws['A4'] = f"Title: {title}"
        ws['A5'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A6'] = "Generated by PropelAI v2.12"
        
        # Detect RFP type (GSA/BPA vs standard UCF)
        section_l_count = len(result.section_l_requirements)
        technical_count = len(result.technical_requirements)
        evaluation_count = len(result.evaluation_requirements)
        
        is_non_ucf_rfp = section_l_count == 0 and (technical_count > 0 or evaluation_count > 0)
        
        # Summary stats
        ws['A8'] = "EXTRACTION SUMMARY"
        ws['A8'].font = Font(bold=True)
        
        stats = result.stats
        row = 9
        ws[f'A{row}'] = "Total Requirements Identified"
        ws[f'B{row}'] = stats.get('total', 0)
        row += 1
        
        ws[f'A{row}'] = "Section L (Submission Instructions)"
        ws[f'B{row}'] = stats.get('section_l', 0)
        row += 1
        
        ws[f'A{row}'] = "Technical Requirements (C/PWS/SOW)"
        ws[f'B{row}'] = stats.get('technical', 0)
        row += 1
        
        ws[f'A{row}'] = "Section M (Evaluation Factors)"
        ws[f'B{row}'] = stats.get('evaluation', 0)
        row += 1
        
        ws[f'A{row}'] = "Attachment Requirements"
        ws[f'B{row}'] = stats.get('attachment', 0)
        row += 2
        
        # Add guidance for non-UCF RFPs (GSA, BPA, etc.)
        if is_non_ucf_rfp:
            ws[f'A{row}'] = "IMPORTANT NOTE"
            ws[f'A{row}'].font = Font(bold=True, color='C65911')  # Orange
            row += 1
            
            ws[f'A{row}'] = "This appears to be a GSA Schedule, BPA, or non-standard RFP format."
            row += 1
            ws[f'A{row}'] = "Section L Compliance is empty because this RFP does not follow"
            row += 1
            ws[f'A{row}'] = "the standard UCF (Uniform Contract Format) with Sections A-M."
            row += 1
            ws[f'A{row}'] = ""
            row += 1
            ws[f'A{row}'] = "For this type of RFP, please review:"
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = f"  • Section M Alignment ({evaluation_count} items) - Submission instructions"
            row += 1
            ws[f'A{row}'] = f"  • Technical Requirements ({technical_count} items) - PWS/SOW requirements"
            row += 2
        
        ws[f'A{row}'] = "BINDING LEVEL BREAKDOWN"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        binding = stats.get('by_binding_level', {})
        ws[f'A{row}'] = "Mandatory (SHALL/MUST)"
        ws[f'B{row}'] = binding.get('mandatory', 0)
        row += 1
        
        ws[f'A{row}'] = "Highly Desirable (SHOULD)"
        ws[f'B{row}'] = binding.get('highly_desirable', 0)
        row += 1
        
        ws[f'A{row}'] = "Desirable (MAY)"
        ws[f'B{row}'] = binding.get('desirable', 0)
        row += 2
        
        # Navigation guide
        ws[f'A{row}'] = "WORKBOOK NAVIGATION"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        sheets_info = [
            ("Section L Compliance", "Submission format and instruction requirements"),
            ("Technical Requirements", "Performance requirements from C/PWS/SOW"),
            ("Section M Alignment", "Evaluation criteria and scoring factors"),
            ("All Requirements", "Complete list of all identified requirements"),
        ]
        
        for sheet_name, description in sheets_info:
            ws[f'A{row}'] = sheet_name
            ws[f'B{row}'] = description
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 55
        ws.column_dimensions['B'].width = 50
    
    def _create_section_l_matrix(self, wb: "Workbook", requirements: List[StructuredRequirement]):
        """
        Create Section L Compliance Matrix.
        
        Per best practices: "Used internally to ensure your volume, headings, 
        page limits, and submission format follow the instructions exactly."
        """
        ws = wb.create_sheet("Section L Compliance")
        
        # Headers per federal recommended model
        headers = [
            "RFP Reference",           # L.4.B.2, etc.
            "Requirement Text",         # VERBATIM - never summarized
            "Source Page",
            "Priority",                 # High/Medium/Low
            "Binding Level",
            "Volume/Section",           # Where to address in proposal
            "Compliance Status",        # Team fills in
            "Compliance Response",      # How we comply
            "Evidence/Notes",
        ]
        
        if not self.include_response_columns:
            headers = headers[:5]  # Only source columns
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                   end_color=self.COLORS['header'], 
                                   fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # If no requirements, add explanatory note for GSA/BPA RFPs
        if not requirements:
            ws.cell(row=2, column=1, value="No Section L requirements found")
            note_cell = ws.cell(row=2, column=2, 
                value="This RFP appears to use a non-standard format (GSA Schedule, BPA, or similar). "
                      "Submission instructions may be located in the 'Section M Alignment' sheet or "
                      "in separate RFP Letter/Instructions documents. Please review the Technical Requirements "
                      "sheet for PWS/SOW requirements and Section M Alignment for evaluation criteria and "
                      "submission instructions.")
            note_cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[2].height = 60
        
        # Write requirements
        for row_num, req in enumerate(requirements, 2):
            # RFP Reference (preserve their numbering!)
            ws.cell(row=row_num, column=1, value=self._safe_cell_value(req.rfp_reference))

            # Full text - VERBATIM (escaped to prevent Excel formula interpretation)
            text_cell = ws.cell(row=row_num, column=2, value=self._safe_cell_value(req.full_text))
            text_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # Page
            ws.cell(row=row_num, column=3, value=req.page_number)
            
            # Priority
            priority = self._get_priority(req.binding_level)
            priority_cell = ws.cell(row=row_num, column=4, value=priority)
            priority_cell.fill = PatternFill(
                start_color=self._get_priority_color(priority),
                end_color=self._get_priority_color(priority),
                fill_type='solid'
            )
            
            # Binding level
            ws.cell(row=row_num, column=5, value=req.binding_level.value)
            
            if self.include_response_columns:
                # Volume/Section (team fills in)
                ws.cell(row=row_num, column=6, value="")
                
                # Compliance Status
                ws.cell(row=row_num, column=7, value="Not Started")
                
                # Response
                ws.cell(row=row_num, column=8, value="")
                
                # Notes
                refs = ", ".join(req.references_to) if req.references_to else ""
                ws.cell(row=row_num, column=9, value=refs)
        
        # Set column widths
        widths = [15, 70, 10, 10, 15, 20, 15, 40, 30]
        for col, width in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Add status dropdown if response columns included
        if self.include_response_columns and len(requirements) > 0:
            status_dv = DataValidation(
                type="list",
                formula1='"Not Started,In Progress,Compliant,Partial,Non-Compliant,N/A"',
                allow_blank=True
            )
            status_dv.add(f"G2:G{len(requirements) + 1}")
            ws.add_data_validation(status_dv)
    
    def _create_technical_matrix(self, wb: "Workbook", requirements: List[StructuredRequirement]):
        """
        Create Technical Requirements Compliance Matrix (C/PWS/SOW).
        
        Per best practices: "Evaluator-facing (if submitted), proving you meet 
        100% of mandatory requirements."
        """
        ws = wb.create_sheet("Technical Requirements")
        
        headers = [
            "Req ID",                   # RFP's own reference
            "Requirement Text",         # VERBATIM
            "Source (PWS/SOW/C)",
            "Page",
            "Priority",                 # High/Medium/Low
            "Binding",
            "Proposal Section",         # Where addressed
            "Compliance Status",
            "Response Strategy",
            "Win Theme",                # How this differentiates us
            "Discriminator/Strength",
            "Proof Point",              # Evidence to cite
            "Evidence Required",
            "Assigned Owner",
            "Interdependencies",        # Related requirements
            "Risk if Non-Compliant",
            "Notes",
        ]
        
        if not self.include_response_columns:
            headers = headers[:6]  # Only source/identification columns
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Write requirements
        for row_num, req in enumerate(requirements, 2):
            ws.cell(row=row_num, column=1, value=self._safe_cell_value(req.rfp_reference))

            text_cell = ws.cell(row=row_num, column=2, value=self._safe_cell_value(req.full_text))
            text_cell.alignment = Alignment(wrap_text=True, vertical='top')

            ws.cell(row=row_num, column=3, value=self._safe_cell_value(req.source_subsection or req.source_section.value))
            ws.cell(row=row_num, column=4, value=req.page_number)
            
            # Priority
            priority = self._get_priority(req.binding_level)
            priority_cell = ws.cell(row=row_num, column=5, value=priority)
            priority_cell.fill = PatternFill(
                start_color=self._get_priority_color(priority),
                end_color=self._get_priority_color(priority),
                fill_type='solid'
            )
            
            ws.cell(row=row_num, column=6, value=req.binding_level.value)
            
            if self.include_response_columns:
                ws.cell(row=row_num, column=7, value="")   # Proposal Section
                ws.cell(row=row_num, column=8, value="Not Started")  # Status
                ws.cell(row=row_num, column=9, value="")   # Response Strategy
                ws.cell(row=row_num, column=10, value="")  # Win Theme
                ws.cell(row=row_num, column=11, value="")  # Discriminator
                ws.cell(row=row_num, column=12, value="")  # Proof Point
                ws.cell(row=row_num, column=13, value="")  # Evidence
                ws.cell(row=row_num, column=14, value="")  # Owner
                ws.cell(row=row_num, column=15, value=", ".join(req.references_to))  # Dependencies
                ws.cell(row=row_num, column=16, value="")  # Risk
                ws.cell(row=row_num, column=17, value="")  # Notes
        
        # Column widths - expand beyond 26 columns using openpyxl's get_column_letter
        from openpyxl.utils import get_column_letter
        widths = [15, 70, 15, 8, 10, 12, 20, 15, 35, 35, 35, 35, 25, 15, 25, 30, 30]
        for col, width in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[get_column_letter(col)].width = width
        
        ws.freeze_panes = 'A2'
        
        if self.include_response_columns and len(requirements) > 0:
            status_dv = DataValidation(
                type="list",
                formula1='"Not Started,In Progress,Fully Compliant,Partial,Non-Compliant,Exceeds,N/A"',
                allow_blank=True
            )
            status_dv.add(f"H2:H{len(requirements) + 1}")
            ws.add_data_validation(status_dv)
    
    def _create_section_m_matrix(self, wb: "Workbook", requirements: List[StructuredRequirement]):
        """
        Create Section M Alignment Matrix.
        
        Per best practices: "Often optional, but powerful. Shows evaluators 
        exactly where you meet or exceed their scoring criteria."
        """
        ws = wb.create_sheet("Section M Alignment")
        
        headers = [
            "Evaluation Factor",        # M.1, M.2, etc.
            "Criterion Text",           # What they will evaluate
            "Page",
            "Weight/Importance",        # If stated
            "Proposal Location",        # Where we address
            "Our Strength",             # What makes us strong here
            "Discriminator",            # How we stand out
            "Proof Points",             # Evidence to cite
            "Risk/Gap",                 # Any concerns
        ]
        
        if not self.include_response_columns:
            headers = headers[:4]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        
        # Write requirements
        for row_num, req in enumerate(requirements, 2):
            ws.cell(row=row_num, column=1, value=self._safe_cell_value(req.rfp_reference))

            text_cell = ws.cell(row=row_num, column=2, value=self._safe_cell_value(req.full_text))
            text_cell.alignment = Alignment(wrap_text=True, vertical='top')

            ws.cell(row=row_num, column=3, value=req.page_number)
            ws.cell(row=row_num, column=4, value="")  # Weight - team determines

            if self.include_response_columns:
                ws.cell(row=row_num, column=5, value="")  # Proposal Location
                ws.cell(row=row_num, column=6, value="")  # Our Strength
                ws.cell(row=row_num, column=7, value="")  # Discriminator
                ws.cell(row=row_num, column=8, value="")  # Proof Points
                ws.cell(row=row_num, column=9, value="")  # Risk/Gap

            # Light yellow for evaluation items
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color=self.COLORS['m_section'],
                    end_color=self.COLORS['m_section'],
                    fill_type='solid'
                )
        
        # Column widths
        widths = [15, 70, 8, 15, 20, 35, 35, 35, 30]
        for col, width in enumerate(widths[:len(headers)], 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        ws.freeze_panes = 'A2'
    
    def _create_all_requirements_sheet(self, wb: "Workbook", requirements: List[StructuredRequirement]):
        """Create sheet with all requirements for reference"""
        ws = wb.create_sheet("All Requirements")
        
        headers = [
            "ID",
            "RFP Reference",
            "Full Text",
            "Category",
            "Section",
            "Binding",
            "Page",
            "Source Document",
            "Cross-References",
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['header'],
                                   end_color=self.COLORS['header'],
                                   fill_type='solid')
        
        for row_num, req in enumerate(requirements, 2):
            ws.cell(row=row_num, column=1, value=self._safe_cell_value(req.generated_id))
            ws.cell(row=row_num, column=2, value=self._safe_cell_value(req.rfp_reference))

            text_cell = ws.cell(row=row_num, column=3, value=self._safe_cell_value(req.full_text))
            text_cell.alignment = Alignment(wrap_text=True, vertical='top')

            ws.cell(row=row_num, column=4, value=req.category.value)
            ws.cell(row=row_num, column=5, value=req.source_section.value)
            ws.cell(row=row_num, column=6, value=req.binding_level.value)
            ws.cell(row=row_num, column=7, value=req.page_number)
            ws.cell(row=row_num, column=8, value=self._safe_cell_value(req.source_document or ""))
            ws.cell(row=row_num, column=9, value=self._safe_cell_value(", ".join(req.references_to)))
            
            # Color by category
            color = self.COLORS.get('att_section')
            if req.category == RequirementCategory.SECTION_L_COMPLIANCE:
                color = self.COLORS['l_section']
            elif req.category == RequirementCategory.TECHNICAL_REQUIREMENT:
                color = self.COLORS['c_section']
            elif req.category == RequirementCategory.EVALUATION_FACTOR:
                color = self.COLORS['m_section']
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col).fill = PatternFill(
                    start_color=color,
                    end_color=color,
                    fill_type='solid'
                )
        
        widths = [15, 15, 70, 15, 10, 15, 8, 25, 30]
        for col, width in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        
        ws.freeze_panes = 'A2'
    
    def _create_cross_reference_sheet(self, wb: "Workbook", result: ExtractionResult):
        """Create cross-reference matrix showing L→M→C linkages"""
        ws = wb.create_sheet("Cross-References")
        
        ws['A1'] = "CROSS-REFERENCE MATRIX"
        ws['A1'].font = Font(size=14, bold=True)
        
        ws['A3'] = "This sheet maps relationships between:"
        ws['A4'] = "• Section L instructions and where to address them"
        ws['A5'] = "• Section M evaluation factors and supporting requirements"
        ws['A6'] = "• Technical requirements and related L/M items"
        
        row = 8
        ws[f'A{row}'] = "DOCUMENT STRUCTURE DETECTED"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        if result.structure:
            for section in result.structure.sections.keys():
                boundary = result.structure.sections[section]
                ws[f'A{row}'] = f"Section {section.value}"
                ws[f'B{row}'] = f"Pages {boundary.start_page}-{boundary.end_page}"
                ws[f'C{row}'] = f"{len(boundary.subsections)} subsections identified"
                row += 1
            
            row += 1
            if result.structure.sow_location:
                ws[f'A{row}'] = "SOW Location"
                ws[f'B{row}'] = result.structure.sow_location
                row += 1
            
            if result.structure.attachments:
                row += 1
                ws[f'A{row}'] = "ATTACHMENTS"
                ws[f'A{row}'].font = Font(bold=True)
                row += 1
                
                for att_id, att_info in result.structure.attachments.items():
                    ws[f'A{row}'] = att_id
                    ws[f'B{row}'] = att_info.document_type
                    ws[f'C{row}'] = f"{att_info.page_count} pages"
                    ws[f'D{row}'] = "Contains requirements" if att_info.contains_requirements else "No requirements"
                    row += 1
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 20


def export_ctm_best_practices(result: ExtractionResult, output_path: str,
                               solicitation_number: str = "", title: str = "") -> str:
    """
    Convenience function to export CTM following best practices.
    
    Usage:
        result = extract_requirements_structured(documents)
        path = export_ctm_best_practices(result, "/path/to/output.xlsx")
    """
    exporter = BestPracticesCTMExporter()
    return exporter.export(result, output_path, solicitation_number, title)
