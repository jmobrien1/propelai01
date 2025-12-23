"""
PropelAI Cycle 5: Excel Exporter - CTM Best Practices Edition
Export compliance matrix to Excel for proposal teams

Creates a professional Compliance Traceability Matrix (CTM) with:
- Proper RFP source tracking (Section, Page, Paragraph)
- Strategic columns (Win Theme, Strength Statement, Proof Points)
- Requirement classification (Mandatory vs Desirable)
- Lifecycle management fields (Status, Owner, Interdependencies)
"""

from typing import Dict, List, Any, Optional, TYPE_CHECKING
from datetime import datetime
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # For type hints when openpyxl not installed

from .models import (
    RequirementNode, RequirementType, ConfidenceLevel,
    ComplianceMatrixRow, ExtractionResult
)


class ExcelExporter:
    """
    Export requirements and compliance matrix to Excel
    
    Creates proposal-team-ready CTM workbooks following best practices:
    - Complete requirement tracking with RFP source
    - Strategic planning columns (Win Theme, Discriminators)
    - Evaluation engineering integration
    - Lifecycle status management
    """
    
    # Color scheme
    COLORS = {
        "header": "1F4E79",           # Dark blue
        "header_strategic": "7B3E19", # Dark brown for strategic columns
        "high_priority": "F4CCCC",    # Light red
        "medium_priority": "FFF2CC",  # Light yellow
        "low_priority": "D9EAD3",     # Light green
        "mandatory": "FADBD8",        # Light red tint for mandatory
        "desirable": "D4EFDF",        # Light green tint for desirable
        "section_c": "CFE2F3",        # Light blue - SOW
        "section_l": "D9D2E9",        # Light purple - Instructions
        "section_m": "FCE5CD",        # Light orange - Evaluation
        "section_pws": "E2EFDA",      # Light green - PWS
    }
    
    # Compliance status options
    STATUS_OPTIONS = [
        "Not Started",
        "In Progress", 
        "Fully Compliant",
        "Partial",
        "Non-Compliant",
        "N/A - Clarification Needed",
        "Needs Clarification"
    ]
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    def export_compliance_matrix(
        self,
        result: ExtractionResult,
        output_path: str,
        solicitation_number: str = "",
        title: str = ""
    ) -> str:
        """
        Export extraction result to Excel workbook
        
        Args:
            result: ExtractionResult from EnhancedComplianceAgent
            output_path: Path for output .xlsx file
            solicitation_number: RFP/solicitation number
            title: Contract title
            
        Returns:
            Path to created file
        """
        wb = Workbook()
        
        # Sheet 1: Summary Dashboard
        self._create_summary_sheet(wb, result, solicitation_number, title)
        
        # Sheet 2: Full Compliance Matrix
        self._create_matrix_sheet(wb, result)
        
        # Sheet 3: High Priority Items
        self._create_priority_sheet(wb, result, "High")
        
        # Sheet 4: By Section (C, L, M breakdown)
        self._create_section_sheet(wb, result)
        
        # Sheet 5: Requirements Graph Data (for advanced users)
        self._create_graph_sheet(wb, result)
        
        # Save workbook
        wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(
        self, 
        wb: Workbook, 
        result: ExtractionResult,
        solicitation_number: str,
        title: str
    ):
        """Create summary dashboard sheet"""
        ws = wb.active
        ws.title = "Summary"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=14)
        header_fill = PatternFill(start_color=self.COLORS["header"], 
                                  end_color=self.COLORS["header"], 
                                  fill_type="solid")
        
        # Title section
        ws.merge_cells('A1:F1')
        ws['A1'] = "PropelAI Compliance Matrix"
        ws['A1'].font = Font(bold=True, size=18)
        
        ws.merge_cells('A2:F2')
        ws['A2'] = f"Solicitation: {solicitation_number}" if solicitation_number else "Solicitation: [Not Specified]"
        ws['A2'].font = Font(size=12)
        
        ws.merge_cells('A3:F3')
        ws['A3'] = f"Title: {title}" if title else ""
        
        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A5'] = ""
        
        # Statistics section
        row = 7
        ws[f'A{row}'] = "EXTRACTION STATISTICS"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        
        # Safely get stats with defaults
        cross_ref_count = getattr(result, 'cross_reference_count', 0) or 0
        extraction_coverage = getattr(result, 'extraction_coverage', 0.0) or 0.0
        duration = getattr(result, 'duration_seconds', 0.0) or 0.0
        stats_dict = getattr(result, 'stats', {}) or {}

        # Calculate valid requirements count from compliance matrix (not req_graph which may be empty)
        total_valid_requirements = 0
        for matrix_row in result.compliance_matrix:
            req_text = getattr(matrix_row, 'requirement_text', '') or ''
            if not self._is_garbage_requirement(req_text):
                total_valid_requirements += 1

        stats = [
            ("Total Requirements", total_valid_requirements),
            ("Cross-References", cross_ref_count),
            ("Documents Processed", stats_dict.get('documents_processed', 0)),
            ("Pages Analyzed", stats_dict.get('total_pages', 0)),
            ("Coverage Estimate", f"{extraction_coverage * 100:.0f}%"),
            ("Processing Time", f"{duration:.1f}s"),
        ]
        
        row += 1
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # By Type breakdown
        row += 2
        ws[f'A{row}'] = "BY REQUIREMENT TYPE"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        for req_type, count in sorted(result.stats.get('by_type', {}).items()):
            ws[f'A{row}'] = req_type.replace('_', ' ').title()
            ws[f'B{row}'] = count
            row += 1
        
        # Priority breakdown
        row += 2
        ws[f'A{row}'] = "BY PRIORITY"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        priority_counts = {"High": 0, "Medium": 0, "Low": 0}
        mandatory_count = 0
        valid_requirements = 0
        
        for matrix_row in result.compliance_matrix:
            req_text = getattr(matrix_row, 'requirement_text', '') or ''
            if self._is_garbage_requirement(req_text):
                continue
            valid_requirements += 1
            priority = getattr(matrix_row, 'priority', 'Medium') or 'Medium'
            priority_counts[priority] = priority_counts.get(priority, 0) + 1
            if self._is_mandatory(req_text):
                mandatory_count += 1
        
        for priority, count in priority_counts.items():
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = count
            
            # Color code
            if priority == "High":
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["high_priority"],
                                                  end_color=self.COLORS["high_priority"],
                                                  fill_type="solid")
            elif priority == "Medium":
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["medium_priority"],
                                                  end_color=self.COLORS["medium_priority"],
                                                  fill_type="solid")
            else:
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["low_priority"],
                                                  end_color=self.COLORS["low_priority"],
                                                  fill_type="solid")
            row += 1
        
        # Mandatory vs Desirable breakdown
        row += 2
        ws[f'A{row}'] = "BY BINDING LEVEL"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 1
        
        ws[f'A{row}'] = "Mandatory (shall/must)"
        ws[f'B{row}'] = mandatory_count
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["mandatory"],
                                          end_color=self.COLORS["mandatory"],
                                          fill_type="solid")
        row += 1
        
        ws[f'A{row}'] = "Desirable (should/may)"
        ws[f'B{row}'] = valid_requirements - mandatory_count
        ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["desirable"],
                                          end_color=self.COLORS["desirable"],
                                          fill_type="solid")
        row += 1
        
        # Valid requirement count note
        row += 2
        ws[f'A{row}'] = f"Valid Requirements (after filtering): {valid_requirements}"
        ws[f'A{row}'].font = Font(italic=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_matrix_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create main compliance matrix sheet with CTM best practice columns"""
        ws = wb.create_sheet("Compliance Matrix")
        
        # CTM Best Practice Headers - organized by function
        # Core Tracking | RFP Source | Classification | Response | Strategic | Lifecycle
        headers = [
            # Core Tracking
            "Req ID",
            "Requirement Text",
            # RFP Source  
            "RFP Section",
            "Page",
            "Source Document",
            # Classification
            "Req Type",
            "Mandatory/Desirable",
            "Priority",
            # Response Planning
            "Proposal Section",
            "Compliance Status",
            "Response Strategy",
            # Strategic (Evaluation Engineering)
            "Win Theme",
            "Discriminator/Strength",
            "Proof Point",
            "Evidence Required",
            # Lifecycle
            "Assigned Owner",
            "Interdependencies",
            "Risk if Non-Compliant",
            "Notes"
        ]
        
        # Header styling
        header_fill = PatternFill(start_color=self.COLORS["header"],
                                  end_color=self.COLORS["header"],
                                  fill_type="solid")
        strategic_fill = PatternFill(start_color=self.COLORS["header_strategic"],
                                     end_color=self.COLORS["header_strategic"],
                                     fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Strategic columns indices (0-based): 11, 12, 13, 14 (Win Theme through Evidence)
        strategic_cols = {12, 13, 14, 15}  # 1-based column numbers
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            if col in strategic_cols:
                cell.fill = strategic_fill
            else:
                cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
            cell.border = thin_border
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Data rows - filter out garbage requirements
        valid_row_idx = 2
        for matrix_row in result.compliance_matrix:
            # Skip garbage requirements (too short, just addresses, etc.)
            req_text = getattr(matrix_row, 'requirement_text', '') or ''
            if self._is_garbage_requirement(req_text):
                continue
            
            # Core tracking
            ws.cell(row=valid_row_idx, column=1, value=getattr(matrix_row, 'requirement_id', ''))
            
            # Truncate long text but keep enough for context
            text = req_text
            if len(text) > 400:
                text = text[:397] + "..."
            ws.cell(row=valid_row_idx, column=2, value=text)
            
            # RFP Source
            section_ref = getattr(matrix_row, 'section_reference', '') or ''
            ws.cell(row=valid_row_idx, column=3, value=section_ref)
            ws.cell(row=valid_row_idx, column=4, value="")  # Page - to be filled
            ws.cell(row=valid_row_idx, column=5, value="")  # Source doc
            
            # Classification
            req_type = getattr(matrix_row, 'requirement_type', 'performance') or 'performance'
            ws.cell(row=valid_row_idx, column=6, value=req_type.replace('_', ' ').title())
            
            # Mandatory vs Desirable based on keywords in text
            mandatory = self._is_mandatory(req_text)
            ws.cell(row=valid_row_idx, column=7, value="Mandatory" if mandatory else "Desirable")
            
            priority = getattr(matrix_row, 'priority', 'Medium') or 'Medium'
            ws.cell(row=valid_row_idx, column=8, value=priority)
            
            # Response Planning
            ws.cell(row=valid_row_idx, column=9, value=getattr(matrix_row, 'proposal_section', '') or '')
            ws.cell(row=valid_row_idx, column=10, value=getattr(matrix_row, 'compliance_status', 'Not Started') or 'Not Started')
            ws.cell(row=valid_row_idx, column=11, value=getattr(matrix_row, 'response_text', '') or '')
            
            # Strategic columns (empty for team to fill)
            ws.cell(row=valid_row_idx, column=12, value="")  # Win Theme
            ws.cell(row=valid_row_idx, column=13, value="")  # Discriminator
            ws.cell(row=valid_row_idx, column=14, value="")  # Proof Point
            evidence = getattr(matrix_row, 'evidence_required', []) or []
            ws.cell(row=valid_row_idx, column=15, value=", ".join(evidence) if isinstance(evidence, list) else str(evidence))
            
            # Lifecycle
            ws.cell(row=valid_row_idx, column=16, value=getattr(matrix_row, 'assigned_owner', '') or '')
            related = getattr(matrix_row, 'related_requirements', []) or []
            ws.cell(row=valid_row_idx, column=17, value=", ".join(related) if isinstance(related, list) else str(related))
            ws.cell(row=valid_row_idx, column=18, value=getattr(matrix_row, 'risk_if_non_compliant', '') or '')
            ws.cell(row=valid_row_idx, column=19, value=getattr(matrix_row, 'notes', '') or '')
            
            # Color coding for priority
            priority_cell = ws.cell(row=valid_row_idx, column=8)
            if priority == "High":
                priority_cell.fill = PatternFill(start_color=self.COLORS["high_priority"],
                                                  end_color=self.COLORS["high_priority"],
                                                  fill_type="solid")
            elif priority == "Medium":
                priority_cell.fill = PatternFill(start_color=self.COLORS["medium_priority"],
                                                  end_color=self.COLORS["medium_priority"],
                                                  fill_type="solid")
            else:
                priority_cell.fill = PatternFill(start_color=self.COLORS["low_priority"],
                                                  end_color=self.COLORS["low_priority"],
                                                  fill_type="solid")
            
            # Color coding for mandatory/desirable
            mand_cell = ws.cell(row=valid_row_idx, column=7)
            if mandatory:
                mand_cell.fill = PatternFill(start_color=self.COLORS["mandatory"],
                                              end_color=self.COLORS["mandatory"],
                                              fill_type="solid")
            else:
                mand_cell.fill = PatternFill(start_color=self.COLORS["desirable"],
                                              end_color=self.COLORS["desirable"],
                                              fill_type="solid")
            
            # Wrap text in requirement column
            ws.cell(row=valid_row_idx, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            
            valid_row_idx += 1
        
        # Add dropdown for Status column (column 10)
        if valid_row_idx > 2:
            dv = DataValidation(type="list", formula1=f'"{",".join(self.STATUS_OPTIONS)}"', allow_blank=True)
            dv.error = "Please select from the list"
            dv.errorTitle = "Invalid Status"
            ws.add_data_validation(dv)
            dv.add(f'J2:J{valid_row_idx - 1}')
            
            # Add dropdown for Mandatory/Desirable (column 7)
            dv2 = DataValidation(type="list", formula1='"Mandatory,Desirable"', allow_blank=True)
            ws.add_data_validation(dv2)
            dv2.add(f'G2:G{valid_row_idx - 1}')
            
            # Add dropdown for Priority (column 8)
            dv3 = DataValidation(type="list", formula1='"High,Medium,Low"', allow_blank=True)
            ws.add_data_validation(dv3)
            dv3.add(f'H2:H{valid_row_idx - 1}')
        
        # Set column widths
        col_widths = {
            'A': 12,   # Req ID
            'B': 60,   # Requirement Text
            'C': 15,   # RFP Section
            'D': 8,    # Page
            'E': 20,   # Source Document
            'F': 15,   # Req Type
            'G': 15,   # Mandatory/Desirable
            'H': 10,   # Priority
            'I': 18,   # Proposal Section
            'J': 15,   # Compliance Status
            'K': 30,   # Response Strategy
            'L': 25,   # Win Theme
            'M': 30,   # Discriminator
            'N': 25,   # Proof Point
            'O': 20,   # Evidence Required
            'P': 15,   # Owner
            'Q': 20,   # Interdependencies
            'R': 20,   # Risk
            'S': 25,   # Notes
        }
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width
    
    def _is_garbage_requirement(self, text: str) -> bool:
        """Filter out garbage extractions that aren't real requirements"""
        if not text:
            return True
        
        text = text.strip()
        
        # Too short to be a real requirement
        if len(text) < 30:
            return True
        
        # Common garbage patterns
        garbage_patterns = [
            r'^[\d\s\-\.\,]+$',  # Just numbers and punctuation
            r'^Page \d+ of \d+',  # Page numbers
            r'^\d+\.\d+ [A-Z]{3,}$',  # Section headers like "2.1 SCOPE"
            r'^[A-Z][a-z]+ \d{1,2}, \d{4}',  # Dates
            r'^https?://',  # URLs
            r'^[\w\.\-]+@[\w\.\-]+',  # Email addresses
            r'^\(\d{3}\) \d{3}-\d{4}',  # Phone numbers
            r'^[A-Z]{2}\s+\d{5}',  # State ZIP like "DC 20593"
            r'^Attachment \d+',  # Attachment headers
            r'^Request for Quote',  # RFQ headers
            r'^Table \d+',  # Table headers
        ]
        
        for pattern in garbage_patterns:
            if re.match(pattern, text, re.IGNORECASE):
                return True
        
        # Address-like content
        if re.search(r'\d+ .+ (Ave|Street|St|Blvd|Road|Rd|Drive|Dr)\.?', text, re.IGNORECASE):
            if len(text) < 100:  # Short address snippet
                return True
        
        # Contains mostly formatting/control characters
        alpha_ratio = sum(1 for c in text if c.isalpha()) / len(text) if text else 0
        if alpha_ratio < 0.4:
            return True
        
        return False
    
    def _is_mandatory(self, text: str) -> bool:
        """Determine if requirement is mandatory based on keywords"""
        mandatory_keywords = [
            r'\bshall\b',
            r'\bmust\b',
            r'\brequired\b',
            r'\bmandatory\b',
            r'\bwill be\b.*\brequired\b',
        ]
        
        text_lower = text.lower()
        for keyword in mandatory_keywords:
            if re.search(keyword, text_lower):
                return True
        return False
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 10
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 40
        ws.column_dimensions['H'].width = 15
        ws.column_dimensions['I'].width = 15
        ws.column_dimensions['J'].width = 30
        ws.column_dimensions['K'].width = 20
        ws.column_dimensions['L'].width = 30
        ws.column_dimensions['M'].width = 30
        
        # Freeze header row
        ws.freeze_panes = 'A2'
    
    def _create_priority_sheet(self, wb: Workbook, result: ExtractionResult, priority: str):
        """Create sheet with priority-filtered items"""
        ws = wb.create_sheet(f"{priority} Priority")
        
        # Filter to priority and exclude garbage
        filtered = []
        for r in result.compliance_matrix:
            req_text = getattr(r, 'requirement_text', '') or ''
            if self._is_garbage_requirement(req_text):
                continue
            row_priority = getattr(r, 'priority', 'Medium') or 'Medium'
            if row_priority == priority:
                filtered.append(r)
        
        # Headers - focused on action items
        headers = [
            "Req ID", "Requirement", "Section", "Mandatory?", 
            "Status", "Owner", "Win Theme", "Proof Point", "Risk"
        ]
        
        header_fill = PatternFill(start_color=self.COLORS["high_priority"] if priority == "High" else self.COLORS["header"],
                                  end_color=self.COLORS["high_priority"] if priority == "High" else self.COLORS["header"],
                                  fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="000000" if priority == "High" else "FFFFFF")
        
        # Freeze header
        ws.freeze_panes = 'A2'
        
        # Data
        for row_idx, matrix_row in enumerate(filtered, 2):
            ws.cell(row=row_idx, column=1, value=getattr(matrix_row, 'requirement_id', ''))
            
            text = getattr(matrix_row, 'requirement_text', '') or ''
            if len(text) > 300:
                text = text[:297] + "..."
            ws.cell(row=row_idx, column=2, value=text)
            
            ws.cell(row=row_idx, column=3, value=getattr(matrix_row, 'section_reference', ''))
            ws.cell(row=row_idx, column=4, value="Yes" if self._is_mandatory(text) else "No")
            ws.cell(row=row_idx, column=5, value=getattr(matrix_row, 'compliance_status', 'Not Started'))
            ws.cell(row=row_idx, column=6, value=getattr(matrix_row, 'assigned_owner', ''))
            ws.cell(row=row_idx, column=7, value="")  # Win Theme - to be filled
            ws.cell(row=row_idx, column=8, value="")  # Proof Point - to be filled
            ws.cell(row=row_idx, column=9, value=getattr(matrix_row, 'risk_if_non_compliant', ''))
            
            # Color code mandatory column
            mand_cell = ws.cell(row=row_idx, column=4)
            if self._is_mandatory(text):
                mand_cell.fill = PatternFill(start_color=self.COLORS["mandatory"],
                                              end_color=self.COLORS["mandatory"],
                                              fill_type="solid")
        
        # Column widths
        ws.column_dimensions['A'].width = 14
        ws.column_dimensions['B'].width = 55
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 25
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 25
    
    def _create_section_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create sheet organized by section"""
        ws = wb.create_sheet("By Section")
        
        # Group by section
        by_section: Dict[str, List] = {}
        for matrix_row in result.compliance_matrix:
            section_ref = getattr(matrix_row, 'section_reference', '') or ''
            section = section_ref[:1] if section_ref else "Other"
            # Map common patterns
            if section_ref.upper().startswith("PWS"):
                section = "PWS"
            elif section_ref.upper().startswith("SOW"):
                section = "C"
            if section not in by_section:
                by_section[section] = []
            by_section[section].append(matrix_row)
        
        row = 1
        
        for section in sorted(by_section.keys()):
            items = by_section[section]
            
            # Filter out garbage
            valid_items = [i for i in items if not self._is_garbage_requirement(
                getattr(i, 'requirement_text', '') or '')]
            
            if not valid_items:
                continue
            
            # Section header
            ws.merge_cells(f'A{row}:E{row}')
            ws[f'A{row}'] = f"Section {section} ({len(valid_items)} requirements)"
            ws[f'A{row}'].font = Font(bold=True, size=14)
            
            # Color by section type
            if section == "C":
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["section_c"],
                                                  end_color=self.COLORS["section_c"],
                                                  fill_type="solid")
            elif section == "L":
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["section_l"],
                                                  end_color=self.COLORS["section_l"],
                                                  fill_type="solid")
            elif section == "M":
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["section_m"],
                                                  end_color=self.COLORS["section_m"],
                                                  fill_type="solid")
            elif section in ["P", "PWS"]:
                ws[f'A{row}'].fill = PatternFill(start_color=self.COLORS["section_pws"],
                                                  end_color=self.COLORS["section_pws"],
                                                  fill_type="solid")
            
            row += 1
            
            # Column headers
            for col, header in enumerate(["ID", "Requirement", "Mandatory?", "Priority", "Status"], 1):
                ws.cell(row=row, column=col, value=header)
                ws.cell(row=row, column=col).font = Font(bold=True)
            row += 1
            
            # Items
            for item in valid_items:
                ws.cell(row=row, column=1, value=getattr(item, 'requirement_id', ''))
                
                text = getattr(item, 'requirement_text', '') or ''
                if len(text) > 200:
                    text = text[:197] + "..."
                ws.cell(row=row, column=2, value=text)
                
                ws.cell(row=row, column=3, value="Yes" if self._is_mandatory(text) else "No")
                ws.cell(row=row, column=4, value=getattr(item, 'priority', 'Medium'))
                ws.cell(row=row, column=5, value=getattr(item, 'compliance_status', 'Not Started'))
                row += 1
            
            row += 1  # Blank row between sections
        
        # Column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
    
    def _create_graph_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create sheet with graph relationship data"""
        ws = wb.create_sheet("Graph Data")
        
        # Headers
        headers = ["ID", "Type", "Section", "Confidence", "References To", "Evaluated By", "Instructed By"]
        
        header_fill = PatternFill(start_color=self.COLORS["header"],
                                  end_color=self.COLORS["header"],
                                  fill_type="solid")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Data
        row = 2
        for req_id, req in result.requirements_graph.items():
            ws.cell(row=row, column=1, value=req_id)
            ws.cell(row=row, column=2, value=req.requirement_type.value)
            ws.cell(row=row, column=3, value=req.source.section_id if req.source else "")
            ws.cell(row=row, column=4, value=req.confidence.value)
            ws.cell(row=row, column=5, value=", ".join(req.references_to[:5]))
            ws.cell(row=row, column=6, value=", ".join(req.evaluated_by[:3]))
            ws.cell(row=row, column=7, value=", ".join(req.instructed_by[:3]))
            row += 1
        
        # Column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        
        ws.freeze_panes = 'A2'


def export_to_excel(
    result: ExtractionResult,
    output_path: str,
    solicitation_number: str = "",
    title: str = ""
) -> str:
    """
    Convenience function to export results to Excel
    
    Args:
        result: ExtractionResult from EnhancedComplianceAgent
        output_path: Path for output .xlsx file
        solicitation_number: RFP/solicitation number
        title: Contract title
        
    Returns:
        Path to created file
    """
    exporter = ExcelExporter()
    return exporter.export_compliance_matrix(result, output_path, solicitation_number, title)
