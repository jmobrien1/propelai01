"""
PropelAI Cycle 5: Excel Traceability Matrix Parser
Parses government RFP requirement matrices (XLSX) into RequirementsGraph

Handles common formats:
- State RTM (Requirements Traceability Matrix)
- Federal compliance matrices
- Vendor response templates
- Multi-sheet workbooks with different requirement types
"""

import os
import re
from typing import List, Dict, Any, Optional, Tuple, Set
from dataclasses import dataclass, field
from enum import Enum
import hashlib

try:
    import openpyxl
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError:
    raise ImportError("openpyxl required: pip install openpyxl")

from .models import RequirementNode, RequirementType, SourceLocation, DocumentType, ConfidenceLevel

# Type alias for requirements graph
RequirementsGraph = Dict[str, RequirementNode]


class ColumnType(Enum):
    """Standard column types in traceability matrices"""
    ID = "id"
    REQUIREMENT = "requirement"
    CATEGORY = "category"
    PRIORITY = "priority"
    MANDATORY = "mandatory"
    SECTION = "section"
    DESCRIPTION = "description"
    RESPONSE = "response"
    COMMENTS = "comments"
    RESPONSIBILITY = "responsibility"
    COMPLIANCE = "compliance"
    EVIDENCE = "evidence"
    REFERENCE = "reference"
    UNKNOWN = "unknown"


@dataclass
class ColumnMapping:
    """Mapping of Excel columns to requirement fields"""
    column_index: int
    column_name: str
    column_type: ColumnType
    confidence: float = 1.0


@dataclass
class SheetConfig:
    """Configuration for a parsed sheet"""
    name: str
    header_row: int
    data_start_row: int
    columns: List[ColumnMapping]
    requirement_count: int = 0
    category_override: Optional[str] = None  # e.g., "Technical", "Business"


@dataclass
class MatrixParseResult:
    """Result of parsing an Excel traceability matrix"""
    requirements_graph: RequirementsGraph
    sheets_parsed: List[SheetConfig]
    total_requirements: int
    by_category: Dict[str, int]
    by_priority: Dict[str, int]
    by_mandatory: Dict[str, int]
    warnings: List[str]
    source_file: str


class ExcelMatrixParser:
    """
    Parse Excel traceability matrices into RequirementsGraph
    
    Features:
    - Auto-detect column mappings
    - Handle multi-sheet workbooks
    - Preserve category/priority/mandatory flags
    - Generate response templates
    """
    
    # Column name patterns for auto-detection
    COLUMN_PATTERNS = {
        ColumnType.ID: [
            r'^id[#\s]*$', r'^req[uirement]*[\s_-]*id', r'^#$', r'^no\.?$',
            r'^item[\s_-]*#?$', r'^ref[erence]*[\s_-]*#?$', r'^row$'
        ],
        ColumnType.REQUIREMENT: [
            r'requirement[\s_-]*(?:text|desc|statement)?',
            r'description', r'^desc\.?$', r'specification',
            r'statement[\s_-]*of[\s_-]*work', r'sow', r'^text$',
            r'service[\s_-]*level[\s_-]*requirement', r'^slr'
        ],
        ColumnType.CATEGORY: [
            r'category', r'type', r'function', r'area', r'domain',
            r'module', r'component', r'group', r'classification'
        ],
        ColumnType.PRIORITY: [
            r'priority', r'importance', r'criticality', r'level',
            r'tier', r'rank'
        ],
        ColumnType.MANDATORY: [
            r'mandatory', r'required', r'must[\s_-]*have',
            r'shall', r'optional', r'desirable', r'm/d', r'm[\s_-]*or[\s_-]*d'
        ],
        ColumnType.SECTION: [
            r'section', r'rfp[\s_-]*section', r'reference', r'source',
            r'paragraph', r'clause'
        ],
        ColumnType.RESPONSE: [
            r'response', r'vendor[\s_-]*response', r'offeror[\s_-]*response',
            r'answer', r'reply', r'met[\s_-]*/?[\s_-]*not[\s_-]*met',
            r'compliance[\s_-]*response'
        ],
        ColumnType.COMMENTS: [
            r'comment', r'note', r'remark', r'observation',
            r'additional[\s_-]*info', r'clarification'
        ],
        ColumnType.RESPONSIBILITY: [
            r'responsibility', r'owner', r'responsible[\s_-]*party',
            r'assigned[\s_-]*to', r'resp\.?[\s_-]*on'
        ],
        ColumnType.COMPLIANCE: [
            r'compliance', r'compliant', r'status', r'met',
            r'satisfied', r'addressed'
        ],
        ColumnType.EVIDENCE: [
            r'evidence', r'proof', r'documentation', r'artifact',
            r'deliverable', r'reference[\s_-]*doc'
        ],
        ColumnType.REFERENCE: [
            r'cross[\s_-]*ref', r'related', r'dependency',
            r'trace[\s_-]*to', r'linked[\s_-]*to'
        ]
    }
    
    # Sheets to skip
    SKIP_SHEETS = [
        'title', 'cover', 'contents', 'toc', 'table of contents',
        'instructions', 'guidance', 'glossary', 'definitions',
        'revision', 'change log', 'history'
    ]
    
    # Priority normalization
    PRIORITY_MAP = {
        'high': 'high', 'h': 'high', '1': 'high', 'critical': 'high',
        'medium': 'medium', 'med': 'medium', 'm': 'medium', '2': 'medium', 'normal': 'medium',
        'low': 'low', 'l': 'low', '3': 'low', 'minor': 'low', 'optional': 'low'
    }
    
    # Mandatory normalization
    MANDATORY_MAP = {
        'mandatory': True, 'required': True, 'must': True, 'shall': True,
        'm': True, 'yes': True, 'y': True, 'true': True, '1': True,
        'desirable': False, 'optional': False, 'should': False, 'may': False,
        'd': False, 'no': False, 'n': False, 'false': False, '0': False
    }
    
    def __init__(self):
        self.warnings: List[str] = []
        self._req_counter = 0
        self._seen_ids: Set[str] = set()
    
    def parse_matrix(
        self,
        file_path: str,
        sheets: Optional[List[str]] = None,
        category_overrides: Optional[Dict[str, str]] = None
    ) -> MatrixParseResult:
        """
        Parse an Excel traceability matrix
        
        Args:
            file_path: Path to Excel file
            sheets: Specific sheets to parse (None = auto-detect)
            category_overrides: Map sheet name to category (e.g., {"5. Technical": "Technical"})
        
        Returns:
            MatrixParseResult with requirements graph and metadata
        """
        self.warnings = []
        self._req_counter = 0
        self._seen_ids = set()
        
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        
        # Load workbook
        wb = openpyxl.load_workbook(file_path, data_only=True)
        
        # Determine sheets to parse
        if sheets:
            sheets_to_parse = [s for s in sheets if s in wb.sheetnames]
        else:
            sheets_to_parse = self._auto_detect_data_sheets(wb)
        
        # Initialize results
        requirements_graph: RequirementsGraph = {}
        sheets_parsed: List[SheetConfig] = []
        by_category: Dict[str, int] = {}
        by_priority: Dict[str, int] = {}
        by_mandatory: Dict[str, int] = {"mandatory": 0, "desirable": 0}
        
        # Parse each sheet
        for sheet_name in sheets_to_parse:
            ws = wb[sheet_name]
            
            # Detect column mappings
            config = self._analyze_sheet(ws, sheet_name)
            
            if not config.columns:
                self.warnings.append(f"No valid columns found in sheet: {sheet_name}")
                continue
            
            # Apply category override if provided
            if category_overrides and sheet_name in category_overrides:
                config.category_override = category_overrides[sheet_name]
            elif not config.category_override:
                # Derive category from sheet name
                config.category_override = self._derive_category(sheet_name)
            
            # Extract requirements
            reqs, req_metadata = self._extract_requirements(ws, config)
            
            # Add to graph
            for req in reqs:
                requirements_graph[req.id] = req
                
                # Update statistics from extracted metadata
                meta = req_metadata.get(req.id, {})
                cat = meta.get("category", "Uncategorized")
                by_category[cat] = by_category.get(cat, 0) + 1
                
                pri = meta.get("priority", "medium")
                by_priority[pri] = by_priority.get(pri, 0) + 1
                
                is_mandatory = meta.get("mandatory", True)
                if is_mandatory:
                    by_mandatory["mandatory"] += 1
                else:
                    by_mandatory["desirable"] += 1
            
            config.requirement_count = len(reqs)
            sheets_parsed.append(config)
        
        wb.close()
        
        return MatrixParseResult(
            requirements_graph=requirements_graph,
            sheets_parsed=sheets_parsed,
            total_requirements=len(requirements_graph),
            by_category=by_category,
            by_priority=by_priority,
            by_mandatory=by_mandatory,
            warnings=self.warnings,
            source_file=os.path.basename(file_path)
        )
    
    def _auto_detect_data_sheets(self, wb: openpyxl.Workbook) -> List[str]:
        """Detect which sheets contain requirement data"""
        data_sheets = []
        
        for sheet_name in wb.sheetnames:
            # Skip known non-data sheets
            name_lower = sheet_name.lower()
            if any(skip in name_lower for skip in self.SKIP_SHEETS):
                continue
            
            ws = wb[sheet_name]
            
            # Check if sheet has data (more than just headers)
            if ws.max_row and ws.max_row > 2:
                # Check for ID-like column in first few columns
                has_id_column = False
                for col in range(1, min(5, ws.max_column + 1)):
                    for row in range(1, min(5, ws.max_row + 1)):
                        val = ws.cell(row=row, column=col).value
                        if val and self._detect_column_type(str(val)) == ColumnType.ID:
                            has_id_column = True
                            break
                    if has_id_column:
                        break
                
                if has_id_column:
                    data_sheets.append(sheet_name)
        
        return data_sheets
    
    def _analyze_sheet(self, ws: Worksheet, sheet_name: str) -> SheetConfig:
        """Analyze sheet structure and detect column mappings"""
        
        # Find header row (look for row with most recognizable column names)
        best_header_row = 1
        best_score = 0
        
        for row in range(1, min(10, ws.max_row + 1)):
            score = 0
            for col in range(1, ws.max_column + 1):
                val = ws.cell(row=row, column=col).value
                if val:
                    col_type = self._detect_column_type(str(val))
                    if col_type != ColumnType.UNKNOWN:
                        score += 1
            
            if score > best_score:
                best_score = score
                best_header_row = row
        
        # Map columns
        columns: List[ColumnMapping] = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=best_header_row, column=col).value
            if val:
                col_name = str(val).strip()
                col_type = self._detect_column_type(col_name)
                columns.append(ColumnMapping(
                    column_index=col,
                    column_name=col_name,
                    column_type=col_type
                ))
        
        return SheetConfig(
            name=sheet_name,
            header_row=best_header_row,
            data_start_row=best_header_row + 1,
            columns=columns
        )
    
    def _detect_column_type(self, column_name: str) -> ColumnType:
        """Detect the type of a column based on its name"""
        name_lower = column_name.lower().strip()
        
        for col_type, patterns in self.COLUMN_PATTERNS.items():
            for pattern in patterns:
                if re.search(pattern, name_lower, re.IGNORECASE):
                    return col_type
        
        return ColumnType.UNKNOWN
    
    def _derive_category(self, sheet_name: str) -> str:
        """Derive a category name from sheet name"""
        # Remove common prefixes like "4. " or "Sheet1 - "
        name = re.sub(r'^\d+[\.\)\s]+', '', sheet_name)
        name = re.sub(r'^Sheet\d+\s*[-:]\s*', '', name, flags=re.IGNORECASE)
        
        # Clean up
        name = name.strip()
        
        # Common mappings
        mappings = {
            'business': 'Business Functions',
            'technical': 'Technical Functions',
            'vendor': 'Vendor Services',
            'performance': 'Performance Requirements',
            'security': 'Security Requirements',
            'functional': 'Functional Requirements',
            'non-functional': 'Non-Functional Requirements',
            'integration': 'Integration Requirements',
            'data': 'Data Requirements',
            'reporting': 'Reporting Requirements'
        }
        
        for key, value in mappings.items():
            if key in name.lower():
                return value
        
        return name or "General"
    
    def _extract_requirements(
        self, 
        ws: Worksheet, 
        config: SheetConfig
    ) -> Tuple[List[RequirementNode], Dict[str, Dict]]:
        """Extract requirements from a worksheet
        
        Returns:
            Tuple of (requirements list, metadata dict keyed by req_id)
        """
        requirements: List[RequirementNode] = []
        metadata_map: Dict[str, Dict] = {}
        
        # Build column lookup
        col_lookup: Dict[ColumnType, int] = {}
        for col_map in config.columns:
            if col_map.column_type != ColumnType.UNKNOWN:
                col_lookup[col_map.column_type] = col_map.column_index
        
        # Need at least an ID or requirement column
        id_col = col_lookup.get(ColumnType.ID)
        req_col = col_lookup.get(ColumnType.REQUIREMENT) or col_lookup.get(ColumnType.DESCRIPTION)
        
        if not id_col and not req_col:
            self.warnings.append(f"No ID or requirement column in {config.name}")
            return requirements, metadata_map
        
        # Extract rows
        for row in range(config.data_start_row, ws.max_row + 1):
            # Get ID
            if id_col:
                raw_id = ws.cell(row=row, column=id_col).value
                if not raw_id or str(raw_id).strip() == "":
                    continue
                req_id = self._normalize_id(str(raw_id).strip(), config.name)
            else:
                self._req_counter += 1
                req_id = f"REQ-{config.name[:3].upper()}-{self._req_counter:04d}"
            
            # Skip duplicates
            if req_id in self._seen_ids:
                continue
            self._seen_ids.add(req_id)
            
            # Get requirement text
            req_text = ""
            if req_col:
                raw_text = ws.cell(row=row, column=req_col).value
                if raw_text:
                    req_text = str(raw_text).strip()
            
            if not req_text:
                # Try to find any text in the row
                for col in range(1, ws.max_column + 1):
                    val = ws.cell(row=row, column=col).value
                    if val and len(str(val)) > 50:
                        req_text = str(val).strip()
                        break
            
            if not req_text:
                continue
            
            # Get category
            category = config.category_override or "General"
            if ColumnType.CATEGORY in col_lookup:
                cat_val = ws.cell(row=row, column=col_lookup[ColumnType.CATEGORY]).value
                if cat_val:
                    category = str(cat_val).strip()
            
            # Get priority
            priority = "medium"
            if ColumnType.PRIORITY in col_lookup:
                pri_val = ws.cell(row=row, column=col_lookup[ColumnType.PRIORITY]).value
                if pri_val:
                    priority = self._normalize_priority(str(pri_val))
            
            # Get mandatory/desirable
            is_mandatory = True
            if ColumnType.MANDATORY in col_lookup:
                mand_val = ws.cell(row=row, column=col_lookup[ColumnType.MANDATORY]).value
                if mand_val:
                    is_mandatory = self._normalize_mandatory(str(mand_val))
            
            # Get section reference
            section = ""
            if ColumnType.SECTION in col_lookup:
                sec_val = ws.cell(row=row, column=col_lookup[ColumnType.SECTION]).value
                if sec_val:
                    section = str(sec_val).strip()
            
            # Determine requirement type
            req_type = self._classify_requirement(req_text, category)
            
            # Build metadata
            metadata = {
                "category": category,
                "mandatory": is_mandatory,
                "source_sheet": config.name,
                "source_row": row
            }
            
            if section:
                metadata["section"] = section
            
            # Store priority in metadata
            metadata["priority"] = priority
            
            # Get response/comments columns for template
            if ColumnType.RESPONSE in col_lookup:
                resp_val = ws.cell(row=row, column=col_lookup[ColumnType.RESPONSE]).value
                if resp_val:
                    metadata["existing_response"] = str(resp_val).strip()
            
            if ColumnType.COMMENTS in col_lookup:
                comm_val = ws.cell(row=row, column=col_lookup[ColumnType.COMMENTS]).value
                if comm_val:
                    metadata["comments"] = str(comm_val).strip()
            
            if ColumnType.RESPONSIBILITY in col_lookup:
                resp_val = ws.cell(row=row, column=col_lookup[ColumnType.RESPONSIBILITY]).value
                if resp_val:
                    metadata["responsibility"] = str(resp_val).strip()
            
            # Create source location
            source = SourceLocation(
                document_name=config.name,
                document_type=DocumentType.ATTACHMENT,
                page_number=row,
                section_id=section or config.name
            )
            
            # Create requirement node
            req = RequirementNode(
                id=req_id,
                text=req_text,
                requirement_type=req_type,
                source=source,
                confidence=ConfidenceLevel.HIGH,  # High confidence from structured source
                keywords=self._extract_keywords(req_text),
                extraction_method="excel_matrix"
            )
            
            # Store metadata separately for statistics
            metadata_map[req_id] = metadata
            
            requirements.append(req)
        
        return requirements, metadata_map
    
    def _normalize_id(self, raw_id: str, sheet_prefix: str) -> str:
        """Normalize requirement ID"""
        # Already has a good format
        if re.match(r'^[A-Z]{2,}-\d+', raw_id):
            return raw_id
        
        # Just a number
        if raw_id.isdigit():
            prefix = sheet_prefix[:3].upper().replace(" ", "")
            return f"{prefix}-{raw_id.zfill(4)}"
        
        # Has letters and numbers
        clean = re.sub(r'[^A-Za-z0-9]', '', raw_id)
        return f"REQ-{clean.upper()}"
    
    def _normalize_priority(self, value: str) -> str:
        """Normalize priority value"""
        val_lower = value.lower().strip()
        return self.PRIORITY_MAP.get(val_lower, "medium")
    
    def _normalize_mandatory(self, value: str) -> bool:
        """Normalize mandatory/desirable value"""
        val_lower = value.lower().strip()
        return self.MANDATORY_MAP.get(val_lower, True)
    
    def _classify_requirement(self, text: str, category: str) -> RequirementType:
        """Classify requirement type from text and category"""
        text_lower = text.lower()
        cat_lower = category.lower()
        
        # Check category first
        if 'performance' in cat_lower:
            return RequirementType.PERFORMANCE_METRIC
        if 'security' in cat_lower:
            return RequirementType.COMPLIANCE
        if 'service' in cat_lower or 'vendor' in cat_lower:
            return RequirementType.PERFORMANCE
        
        # Check text patterns
        if any(w in text_lower for w in ['shall not', 'must not', 'prohibited', 'forbidden']):
            return RequirementType.PROHIBITION
        if any(w in text_lower for w in ['deliver', 'provide', 'submit', 'produce']):
            return RequirementType.DELIVERABLE
        if any(w in text_lower for w in ['comply', 'accordance', 'conform', 'adhere']):
            return RequirementType.COMPLIANCE
        if any(w in text_lower for w in ['qualified', 'certified', 'licensed', 'experience']):
            return RequirementType.QUALIFICATION
        if any(w in text_lower for w in ['evaluate', 'score', 'assess', 'criteria']):
            return RequirementType.EVALUATION_CRITERION
        
        return RequirementType.PERFORMANCE
    
    def _extract_keywords(self, text: str) -> List[str]:
        """Extract keywords from requirement text"""
        # Simple keyword extraction - significant words
        import re
        words = re.findall(r'\b[a-z]{4,}\b', text.lower())
        
        # Filter common stop words
        stop_words = {
            'shall', 'must', 'will', 'should', 'that', 'this', 'with', 'from',
            'have', 'been', 'were', 'being', 'which', 'their', 'there', 'these',
            'those', 'about', 'would', 'could', 'other', 'after', 'before'
        }
        
        keywords = [w for w in words if w not in stop_words]
        
        # Return unique, limited to 10
        seen = set()
        unique = []
        for kw in keywords:
            if kw not in seen:
                seen.add(kw)
                unique.append(kw)
                if len(unique) >= 10:
                    break
        
        return unique
    
    def generate_response_template(
        self,
        result: MatrixParseResult,
        output_path: str,
        include_guidance: bool = True
    ) -> str:
        """
        Generate a response template Excel file
        
        Args:
            result: Parse result with requirements
            output_path: Path for output file
            include_guidance: Add guidance column with tips
        
        Returns:
            Path to generated file
        """
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        default_sheet = wb.active
        wb.remove(default_sheet)
        
        # Group requirements by category (use source document name)
        by_category: Dict[str, List[RequirementNode]] = {}
        for req in result.requirements_graph.values():
            # Use source document name as category, or derive from by_category in result
            if req.source:
                cat = req.source.document_name
            else:
                cat = "General"
            if cat not in by_category:
                by_category[cat] = []
            by_category[cat].append(req)
        
        # Create sheet per category
        for category, reqs in sorted(by_category.items()):
            # Clean sheet name (max 31 chars, no special chars)
            sheet_name = re.sub(r'[^\w\s]', '', category)[:31]
            ws = wb.create_sheet(title=sheet_name)
            
            # Headers
            headers = [
                "ID", "Requirement", "Type",
                "Compliance Status", "Vendor Response", "Evidence/Reference"
            ]
            if include_guidance:
                headers.append("Response Guidance")
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = openpyxl.styles.Font(bold=True)
                cell.fill = openpyxl.styles.PatternFill(
                    start_color="1F4E79", end_color="1F4E79", fill_type="solid"
                )
                cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            
            # Data rows
            for row_idx, req in enumerate(reqs, 2):
                ws.cell(row=row_idx, column=1, value=req.id)
                ws.cell(row=row_idx, column=2, value=req.text)
                ws.cell(row=row_idx, column=3, value=req.requirement_type.value)
                ws.cell(row=row_idx, column=4, value="")  # Compliance status - to fill
                ws.cell(row=row_idx, column=5, value="")  # Response - to fill
                ws.cell(row=row_idx, column=6, value="")  # Evidence - to fill
                
                if include_guidance:
                    guidance = self._generate_guidance(req)
                    ws.cell(row=row_idx, column=7, value=guidance)
            
            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 60
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 50
            ws.column_dimensions['F'].width = 30
            if include_guidance:
                ws.column_dimensions['G'].width = 40
            
            # Freeze header row
            ws.freeze_panes = 'A2'
        
        # Create summary sheet
        summary = wb.create_sheet(title="Summary", index=0)
        summary.cell(row=1, column=1, value="Category")
        summary.cell(row=1, column=2, value="Total")
        summary.cell(row=1, column=3, value="Mandatory")
        summary.cell(row=1, column=4, value="Desirable")
        
        for col in range(1, 5):
            summary.cell(row=1, column=col).font = openpyxl.styles.Font(bold=True)
        
        row = 2
        for category, reqs in sorted(by_category.items()):
            summary.cell(row=row, column=1, value=category)
            summary.cell(row=row, column=2, value=len(reqs))
            # Use result.by_mandatory for accurate counts
            summary.cell(row=row, column=3, value="-")
            summary.cell(row=row, column=4, value="-")
            row += 1
        
        # Totals row
        summary.cell(row=row, column=1, value="TOTAL")
        summary.cell(row=row, column=2, value=result.total_requirements)
        summary.cell(row=row, column=3, value=result.by_mandatory.get("mandatory", 0))
        summary.cell(row=row, column=4, value=result.by_mandatory.get("desirable", 0))
        for col in range(1, 5):
            summary.cell(row=row, column=col).font = openpyxl.styles.Font(bold=True)
        
        wb.save(output_path)
        return output_path
    
    def _generate_guidance(self, req: RequirementNode) -> str:
        """Generate response guidance for a requirement"""
        text_lower = req.text.lower()
        
        # Check for specific patterns
        if 'shall' in text_lower or 'must' in text_lower:
            return "Mandatory requirement - provide explicit compliance statement"
        if 'experience' in text_lower or 'years' in text_lower:
            return "Reference specific past performance and team qualifications"
        if 'security' in text_lower or 'compliance' in text_lower:
            return "Cite specific certifications, standards, and security controls"
        if 'report' in text_lower or 'dashboard' in text_lower:
            return "Describe reporting capabilities with examples or screenshots"
        if 'integrate' in text_lower or 'interface' in text_lower:
            return "Detail integration approach, APIs, and technical methodology"
        if 'support' in text_lower or 'maintain' in text_lower:
            return "Outline support model, SLAs, and escalation procedures"
        if 'train' in text_lower:
            return "Describe training approach, materials, and delivery method"
        
        return "Address requirement with specific approach and evidence"


# Convenience function
def parse_excel_matrix(
    file_path: str,
    sheets: Optional[List[str]] = None,
    category_overrides: Optional[Dict[str, str]] = None
) -> MatrixParseResult:
    """
    Parse an Excel traceability matrix
    
    Args:
        file_path: Path to Excel file
        sheets: Specific sheets to parse (None = auto-detect)
        category_overrides: Map sheet name to category
    
    Returns:
        MatrixParseResult with requirements and metadata
    """
    parser = ExcelMatrixParser()
    return parser.parse_matrix(file_path, sheets, category_overrides)
