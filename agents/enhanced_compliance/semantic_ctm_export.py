"""
PropelAI v2.8: Semantic CTM Exporter

Creates professional Compliance Traceability Matrix from semantic extraction.
Follows CTM Best Practices:
- Clean, readable requirement text
- Proper RFP section mapping (L, M, C, PWS, SOW)
- Strategic columns for proposal development
- Evaluation engineering integration
"""

from typing import Dict, List, Any, Optional
from datetime import datetime
import os
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    Workbook = None  # For type hints when openpyxl not installed

from .semantic_extractor import (
    ExtractedRequirement, ExtractionResult, 
    SemanticRequirementType, RFPSection
)


class SemanticCTMExporter:
    """
    Export semantically extracted requirements to professional CTM.
    
    Output Structure (19 columns):
    1. Req ID
    2. Requirement Text (clean, readable)
    3. RFP Section (L, M, C, PWS, etc.)
    4. Section Reference (L.4.B.2, PWS 2.3, etc.)
    5. Page
    6. Source Document
    7. Requirement Type
    8. Mandatory/Desirable
    9. Priority
    10. Proposal Section
    11. Compliance Status
    12. Response Strategy
    13. Win Theme [Strategic]
    14. Discriminator/Strength [Strategic]
    15. Proof Point [Strategic]
    16. Evidence Required
    17. Assigned Owner
    18. Interdependencies
    19. Notes
    """
    
    # Color scheme
    COLORS = {
        "header": "1F4E79",              # Dark blue
        "header_strategic": "7B3E19",    # Dark brown for strategic columns
        "high_priority": "F4CCCC",       # Light red
        "medium_priority": "FFF2CC",     # Light yellow
        "low_priority": "D9EAD3",        # Light green
        "mandatory": "FADBD8",           # Light red tint
        "desirable": "D4EFDF",           # Light green tint
        "section_l": "D9D2E9",           # Light purple
        "section_m": "FCE5CD",           # Light orange
        "section_c": "CFE2F3",           # Light blue
        "section_pws": "E2EFDA",         # Light green
        "performance": "E3F2FD",         # Very light blue
        "proposal": "F3E5F5",            # Very light purple
        "evaluation": "FFF3E0",          # Very light orange
    }
    
    STATUS_OPTIONS = [
        "Not Started",
        "In Progress", 
        "Fully Compliant",
        "Partial",
        "Non-Compliant",
        "N/A",
        "Needs Clarification"
    ]
    
    PRIORITY_OPTIONS = ["HIGH", "MEDIUM", "LOW"]
    BINDING_OPTIONS = ["Mandatory", "Desirable"]
    
    # Column widths
    COLUMN_WIDTHS = {
        'A': 14,   # Req ID
        'B': 65,   # Requirement Text
        'C': 10,   # RFP Section
        'D': 14,   # Section Ref
        'E': 8,    # Page
        'F': 25,   # Source Doc
        'G': 18,   # Req Type
        'H': 14,   # Mandatory
        'I': 10,   # Priority
        'J': 18,   # Proposal Section
        'K': 16,   # Status
        'L': 25,   # Response Strategy
        'M': 25,   # Win Theme
        'N': 30,   # Discriminator
        'O': 25,   # Proof Point
        'P': 20,   # Evidence
        'Q': 15,   # Owner
        'R': 20,   # Interdependencies
        'S': 25,   # Notes
    }
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required: pip install openpyxl")
        
        self._thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def export(
        self,
        result: ExtractionResult,
        output_path: str,
        solicitation_number: str = "",
        title: str = ""
    ) -> str:
        """Export extraction result to professional CTM Excel workbook"""
        wb = Workbook()
        
        # Create sheets
        self._create_summary_sheet(wb, result, solicitation_number, title)
        self._create_ctm_sheet(wb, result)
        self._create_by_type_sheet(wb, result)
        self._create_by_section_sheet(wb, result)
        self._create_high_priority_sheet(wb, result)
        
        # Save
        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
        wb.save(output_path)
        
        return output_path
    
    def _create_summary_sheet(
        self, 
        wb: Workbook, 
        result: ExtractionResult,
        solicitation_number: str,
        title: str
    ):
        """Create summary dashboard sheet"""
        ws = wb.active
        ws.title = "Summary"
        
        # Header
        ws['A1'] = "PropelAI Compliance Traceability Matrix"
        ws['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws.merge_cells('A1:C1')
        
        # Solicitation info
        ws['A3'] = f"Solicitation: {solicitation_number or '[Not Specified]'}"
        ws['A4'] = f"Title: {title or '[Not Specified]'}"
        ws['A5'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # Stats
        ws['A7'] = "EXTRACTION STATISTICS"
        ws['A7'].font = Font(bold=True, size=12)
        
        row = 8
        stats = result.stats
        
        ws[f'A{row}'] = "Total Requirements"
        ws[f'B{row}'] = stats.get('total', len(result.requirements))
        row += 1
        
        ws[f'A{row}'] = "Mandatory"
        ws[f'B{row}'] = stats.get('mandatory', 0)
        ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['mandatory'], 
                                          end_color=self.COLORS['mandatory'], fill_type='solid')
        row += 1
        
        ws[f'A{row}'] = "Desirable"
        ws[f'B{row}'] = stats.get('desirable', 0)
        ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['desirable'], 
                                          end_color=self.COLORS['desirable'], fill_type='solid')
        row += 2
        
        # By Type
        ws[f'A{row}'] = "BY REQUIREMENT TYPE"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        type_stats = stats.get('by_type', {})
        for type_name, count in sorted(type_stats.items(), key=lambda x: -x[1]):
            ws[f'A{row}'] = self._format_type_name(type_name)
            ws[f'B{row}'] = count
            row += 1
        
        row += 1
        
        # By Section
        ws[f'A{row}'] = "BY RFP SECTION"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        section_stats = stats.get('by_section', {})
        for section_name, count in sorted(section_stats.items(), key=lambda x: -x[1]):
            ws[f'A{row}'] = self._format_section_name(section_name)
            ws[f'B{row}'] = count
            row += 1
        
        row += 1
        
        # By Priority
        ws[f'A{row}'] = "BY PRIORITY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        priority_stats = stats.get('by_priority', {})
        for priority in ['HIGH', 'MEDIUM', 'LOW']:
            count = priority_stats.get(priority, 0)
            if count > 0:
                ws[f'A{row}'] = priority
                ws[f'B{row}'] = count
                color_key = f"{priority.lower()}_priority"
                if color_key in self.COLORS:
                    ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS[color_key], 
                                                      end_color=self.COLORS[color_key], fill_type='solid')
                row += 1
        
        # Warnings
        if result.warnings:
            row += 1
            ws[f'A{row}'] = "WARNINGS"
            ws[f'A{row}'].font = Font(bold=True, size=12, color="CC0000")
            row += 1
            for warning in result.warnings[:10]:
                ws[f'A{row}'] = warning
                row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    def _create_ctm_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create the main Compliance Traceability Matrix sheet"""
        ws = wb.create_sheet("Compliance Matrix")
        
        # Headers (19 columns)
        headers = [
            ("Req ID", "header"),
            ("Requirement Text", "header"),
            ("RFP Section", "header"),
            ("Section Ref", "header"),
            ("Page", "header"),
            ("Source Document", "header"),
            ("Req Type", "header"),
            ("Mandatory/Desirable", "header"),
            ("Priority", "header"),
            ("Proposal Section", "header"),
            ("Compliance Status", "header"),
            ("Response Strategy", "header"),
            ("Win Theme", "header_strategic"),
            ("Discriminator/Strength", "header_strategic"),
            ("Proof Point", "header_strategic"),
            ("Evidence Required", "header"),
            ("Assigned Owner", "header"),
            ("Interdependencies", "header"),
            ("Notes", "header"),
        ]
        
        # Write headers
        for col, (header_text, color_key) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header_text)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLORS[color_key], 
                                    end_color=self.COLORS[color_key], fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = self._thin_border
        
        # Write data
        row = 2
        for req in result.requirements:
            # Column A: Req ID
            ws.cell(row=row, column=1, value=req.id)
            
            # Column B: Requirement Text (clean)
            ws.cell(row=row, column=2, value=req.text)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Column C: RFP Section
            ws.cell(row=row, column=3, value=req.rfp_section.value)
            
            # Column D: Section Reference
            ws.cell(row=row, column=4, value=req.section_reference)
            
            # Column E: Page
            ws.cell(row=row, column=5, value=req.page_number if req.page_number > 0 else "")
            
            # Column F: Source Document
            ws.cell(row=row, column=6, value=self._truncate_filename(req.source_document))
            
            # Column G: Requirement Type
            ws.cell(row=row, column=7, value=self._format_type_name(req.requirement_type.value))
            
            # Column H: Mandatory/Desirable
            binding = "Mandatory" if req.is_mandatory else "Desirable"
            ws.cell(row=row, column=8, value=binding)
            color = self.COLORS['mandatory'] if req.is_mandatory else self.COLORS['desirable']
            ws.cell(row=row, column=8).fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            
            # Column I: Priority
            ws.cell(row=row, column=9, value=req.priority)
            priority_color = self.COLORS.get(f"{req.priority.lower()}_priority")
            if priority_color:
                ws.cell(row=row, column=9).fill = PatternFill(start_color=priority_color, 
                                                               end_color=priority_color, fill_type='solid')
            
            # Column J: Proposal Section (empty for team to fill)
            ws.cell(row=row, column=10, value="")
            
            # Column K: Compliance Status
            ws.cell(row=row, column=11, value="Not Started")
            
            # Column L: Response Strategy (empty)
            ws.cell(row=row, column=12, value="")
            
            # Columns M-O: Strategic columns (empty for team)
            ws.cell(row=row, column=13, value="")  # Win Theme
            ws.cell(row=row, column=14, value="")  # Discriminator
            ws.cell(row=row, column=15, value="")  # Proof Point
            
            # Column P: Evidence Required
            evidence = ", ".join(req.constraints) if req.constraints else ""
            ws.cell(row=row, column=16, value=evidence)
            
            # Column Q: Assigned Owner (empty)
            ws.cell(row=row, column=17, value="")
            
            # Column R: Interdependencies
            interdeps = []
            if req.references_sections:
                interdeps.extend([f"Sec {s}" for s in req.references_sections])
            if req.references_attachments:
                interdeps.extend(req.references_attachments)
            ws.cell(row=row, column=18, value=", ".join(interdeps))
            
            # Column S: Notes
            notes = []
            if req.action_verb:
                notes.append(f"Action: {req.action_verb}")
            if req.related_evaluation_factor:
                notes.append(f"Eval: {req.related_evaluation_factor}")
            ws.cell(row=row, column=19, value="; ".join(notes))
            
            # Apply borders
            for col in range(1, 20):
                ws.cell(row=row, column=col).border = self._thin_border
            
            row += 1
        
        # Column widths
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add dropdowns
        self._add_dropdowns(ws, row - 1)
    
    def _create_by_type_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create sheet organized by requirement type"""
        ws = wb.create_sheet("By Type")
        
        # Group by type
        by_type = {}
        for req in result.requirements:
            type_name = req.requirement_type.value
            if type_name not in by_type:
                by_type[type_name] = []
            by_type[type_name].append(req)
        
        row = 1
        for type_name in sorted(by_type.keys()):
            reqs = by_type[type_name]
            
            # Section header
            ws.cell(row=row, column=1, value=f"{self._format_type_name(type_name)} ({len(reqs)})")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Column headers
            headers = ["ID", "Requirement", "Section", "Mandatory", "Priority"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color=self.COLORS['header'], 
                                        end_color=self.COLORS['header'], fill_type='solid')
                cell.font = Font(bold=True, color="FFFFFF")
            row += 1
            
            # Requirements
            for req in reqs:
                ws.cell(row=row, column=1, value=req.id)
                ws.cell(row=row, column=2, value=req.text[:200] + "..." if len(req.text) > 200 else req.text)
                ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=3, value=req.section_reference or req.rfp_section.value)
                ws.cell(row=row, column=4, value="Mandatory" if req.is_mandatory else "Desirable")
                ws.cell(row=row, column=5, value=req.priority)
                row += 1
            
            row += 1  # Blank row between sections
        
        # Column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 10
    
    def _create_by_section_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create sheet organized by RFP section"""
        ws = wb.create_sheet("By Section")
        
        # Group by section
        by_section = {}
        for req in result.requirements:
            section = req.rfp_section.value
            if section not in by_section:
                by_section[section] = []
            by_section[section].append(req)
        
        # Sort sections in logical order
        section_order = ['L', 'M', 'C', 'PWS', 'SOW', 'F', 'H', 'J', 'ATT', 'UNK']
        sorted_sections = sorted(by_section.keys(), 
                                  key=lambda x: section_order.index(x) if x in section_order else 99)
        
        row = 1
        for section in sorted_sections:
            reqs = by_section[section]
            
            # Section header
            section_name = self._format_section_name(section)
            ws.cell(row=row, column=1, value=f"{section_name} ({len(reqs)} requirements)")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="1F4E79")
            ws.merge_cells(f'A{row}:E{row}')
            row += 1
            
            # Column headers
            headers = ["ID", "Requirement", "Reference", "Type", "Priority"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                
                # Color by section
                color = self.COLORS.get(f"section_{section.lower()}", self.COLORS['header'])
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            row += 1
            
            # Requirements
            for req in sorted(reqs, key=lambda r: r.section_reference or ""):
                ws.cell(row=row, column=1, value=req.id)
                ws.cell(row=row, column=2, value=req.text[:200] + "..." if len(req.text) > 200 else req.text)
                ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
                ws.cell(row=row, column=3, value=req.section_reference)
                ws.cell(row=row, column=4, value=self._format_type_name(req.requirement_type.value))
                ws.cell(row=row, column=5, value=req.priority)
                row += 1
            
            row += 1  # Blank row
        
        # Column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 10
    
    def _create_high_priority_sheet(self, wb: Workbook, result: ExtractionResult):
        """Create sheet with high priority items for immediate attention"""
        ws = wb.create_sheet("High Priority")
        
        # Filter high priority
        high_priority = [r for r in result.requirements if r.priority == "HIGH"]
        
        # Sort by mandatory first, then by type
        high_priority.sort(key=lambda r: (not r.is_mandatory, r.requirement_type.value))
        
        # Headers
        headers = [
            "Req ID", "Requirement", "Section", "Type", 
            "Mandatory", "Status", "Owner", "Win Theme", "Risk if Missed"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color=self.COLORS['high_priority'], 
                                    end_color=self.COLORS['high_priority'], fill_type='solid')
            cell.fill = PatternFill(start_color="CC0000", end_color="CC0000", fill_type='solid')
            cell.border = self._thin_border
        
        # Data
        row = 2
        for req in high_priority:
            ws.cell(row=row, column=1, value=req.id)
            ws.cell(row=row, column=2, value=req.text[:300] + "..." if len(req.text) > 300 else req.text)
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            ws.cell(row=row, column=3, value=req.section_reference or req.rfp_section.value)
            ws.cell(row=row, column=4, value=self._format_type_name(req.requirement_type.value))
            ws.cell(row=row, column=5, value="Mandatory" if req.is_mandatory else "Desirable")
            ws.cell(row=row, column=6, value="Not Started")
            ws.cell(row=row, column=7, value="")  # Owner
            ws.cell(row=row, column=8, value="")  # Win Theme
            ws.cell(row=row, column=9, value="")  # Risk
            
            # Borders
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self._thin_border
            
            row += 1
        
        # Column widths
        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 70
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 15
        ws.column_dimensions['H'].width = 25
        ws.column_dimensions['I'].width = 25
        
        # Freeze header
        ws.freeze_panes = 'A2'
    
    def _add_dropdowns(self, ws, max_row: int):
        """Add dropdown validations to the CTM sheet"""
        # Status dropdown (column K)
        status_dv = DataValidation(
            type="list",
            formula1='"' + ','.join(self.STATUS_OPTIONS) + '"',
            allow_blank=True
        )
        status_dv.error = "Please select from the list"
        status_dv.errorTitle = "Invalid Status"
        ws.add_data_validation(status_dv)
        status_dv.add(f'K2:K{max_row + 1}')
        
        # Binding dropdown (column H)
        binding_dv = DataValidation(
            type="list",
            formula1='"' + ','.join(self.BINDING_OPTIONS) + '"',
            allow_blank=True
        )
        ws.add_data_validation(binding_dv)
        binding_dv.add(f'H2:H{max_row + 1}')
        
        # Priority dropdown (column I)
        priority_dv = DataValidation(
            type="list",
            formula1='"' + ','.join(self.PRIORITY_OPTIONS) + '"',
            allow_blank=True
        )
        ws.add_data_validation(priority_dv)
        priority_dv.add(f'I2:I{max_row + 1}')
    
    def _format_type_name(self, type_value: str) -> str:
        """Format requirement type for display"""
        type_names = {
            'PERFORMANCE_REQ': 'Performance Requirement',
            'PERFORMANCE_REQUIREMENT': 'Performance Requirement',
            'PROPOSAL_INSTRUCTION': 'Proposal Instruction',
            'EVALUATION_CRITERION': 'Evaluation Criterion',
            'DELIVERABLE': 'Deliverable',
            'QUALIFICATION': 'Qualification',
            'COMPLIANCE_CLAUSE': 'Compliance/Clause',
            'PROHIBITION': 'Prohibition',
            'OTHER': 'Other',
        }
        return type_names.get(type_value, type_value.replace('_', ' ').title())
    
    def _format_section_name(self, section_value: str) -> str:
        """Format RFP section for display"""
        section_names = {
            'L': 'Section L - Instructions to Offerors',
            'M': 'Section M - Evaluation Factors',
            'C': 'Section C - Description/SOW',
            'PWS': 'Performance Work Statement',
            'SOW': 'Statement of Work',
            'F': 'Section F - Deliveries/Performance',
            'H': 'Section H - Special Requirements',
            'J': 'Section J - Attachments',
            'ATT': 'Attachments',
            'UNK': 'Unspecified Section',
        }
        return section_names.get(section_value, f"Section {section_value}")
    
    def _truncate_filename(self, filename: str, max_len: int = 40) -> str:
        """Truncate long filenames"""
        if len(filename) <= max_len:
            return filename
        return filename[:max_len-3] + "..."
