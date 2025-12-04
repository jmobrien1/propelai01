"""
J.P-1 Matrix Parser
===================

Parses the GSA OASIS+ Domain Qualifications Matrix (Attachment J.P-1)
to dynamically load scoring rules.

The J.P-1 is an Excel workbook with tabs for each domain containing
the scoring criteria, point values, and validation rules.
"""

import re
import logging
from pathlib import Path
from typing import List, Dict, Optional, Any, Tuple
from decimal import Decimal

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.worksheet.worksheet import Worksheet
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from .models import (
    OASISDomain,
    ScoringCriteria,
    DomainType,
    CriteriaType,
    BusinessSize,
)

logger = logging.getLogger(__name__)


# Domain tab name mappings (J.P-1 sheet names to DomainType)
DOMAIN_TAB_MAPPINGS = {
    "technical": DomainType.TECHNICAL_ENGINEERING,
    "tech": DomainType.TECHNICAL_ENGINEERING,
    "t&e": DomainType.TECHNICAL_ENGINEERING,
    "technical & engineering": DomainType.TECHNICAL_ENGINEERING,
    "management": DomainType.MANAGEMENT_ADVISORY,
    "mgmt": DomainType.MANAGEMENT_ADVISORY,
    "m&a": DomainType.MANAGEMENT_ADVISORY,
    "management & advisory": DomainType.MANAGEMENT_ADVISORY,
    "enterprise": DomainType.ENTERPRISE_SOLUTIONS,
    "es": DomainType.ENTERPRISE_SOLUTIONS,
    "enterprise solutions": DomainType.ENTERPRISE_SOLUTIONS,
    "intelligence": DomainType.INTELLIGENCE_SERVICES,
    "intel": DomainType.INTELLIGENCE_SERVICES,
    "logistics": DomainType.LOGISTICS,
    "log": DomainType.LOGISTICS,
    "environmental": DomainType.ENVIRONMENTAL,
    "env": DomainType.ENVIRONMENTAL,
    "facilities": DomainType.FACILITIES,
    "fac": DomainType.FACILITIES,
    "r&d": DomainType.RESEARCH_DEVELOPMENT,
    "research": DomainType.RESEARCH_DEVELOPMENT,
}

# Common column header patterns
CRITERIA_ID_PATTERNS = [
    r"^l\.\d+",  # L.5.2.1
    r"^criteria",
    r"^ref",
    r"^id",
    r"^number",
]

DESCRIPTION_PATTERNS = [
    r"description",
    r"requirement",
    r"criteria\s*text",
    r"qualification",
]

POINTS_PATTERNS = [
    r"points?",
    r"score",
    r"max",
    r"value",
]


class JP1MatrixParser:
    """
    Parses OASIS+ J.P-1 Qualifications Matrix Excel files.

    Extracts scoring criteria for each domain to create a dynamic
    rules engine that adapts to GSA amendments.
    """

    def __init__(self):
        self.domains: Dict[DomainType, OASISDomain] = {}
        self._raw_data: Dict[str, Any] = {}

    def parse_file(self, filepath: str) -> Dict[DomainType, OASISDomain]:
        """
        Parse a J.P-1 Excel file and extract all domains.

        Args:
            filepath: Path to the J.P-1 Excel file

        Returns:
            Dictionary mapping DomainType to OASISDomain objects
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for J.P-1 parsing")

        path = Path(filepath)
        if not path.exists():
            raise FileNotFoundError(f"J.P-1 file not found: {filepath}")

        logger.info(f"Parsing J.P-1 Matrix: {filepath}")

        wb = load_workbook(filepath, read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            domain_type = self._detect_domain_type(sheet_name)
            if domain_type:
                logger.info(f"Parsing domain sheet: {sheet_name} -> {domain_type.value}")
                domain = self._parse_domain_sheet(wb[sheet_name], domain_type, sheet_name)
                self.domains[domain_type] = domain

        wb.close()

        logger.info(f"Parsed {len(self.domains)} domains with "
                   f"{sum(len(d.criteria) for d in self.domains.values())} total criteria")

        return self.domains

    def _detect_domain_type(self, sheet_name: str) -> Optional[DomainType]:
        """Detect domain type from sheet name"""
        normalized = sheet_name.lower().strip()

        # Direct mapping
        if normalized in DOMAIN_TAB_MAPPINGS:
            return DOMAIN_TAB_MAPPINGS[normalized]

        # Fuzzy matching
        for key, domain_type in DOMAIN_TAB_MAPPINGS.items():
            if key in normalized:
                return domain_type

        return None

    def _parse_domain_sheet(
        self,
        sheet: Worksheet,
        domain_type: DomainType,
        sheet_name: str
    ) -> OASISDomain:
        """Parse a single domain sheet from the workbook"""

        # Find header row and column mappings
        header_row, col_map = self._find_headers(sheet)

        if not col_map:
            logger.warning(f"Could not find headers in sheet: {sheet_name}")
            return OASISDomain(
                domain_type=domain_type,
                name=sheet_name,
                description=f"OASIS+ {sheet_name} Domain"
            )

        # Extract criteria rows
        criteria_list = []
        for row_idx in range(header_row + 1, sheet.max_row + 1):
            criteria = self._parse_criteria_row(sheet, row_idx, col_map, domain_type)
            if criteria:
                criteria_list.append(criteria)

        # Extract thresholds if present
        thresholds = self._extract_thresholds(sheet)

        domain = OASISDomain(
            domain_type=domain_type,
            name=self._format_domain_name(domain_type),
            description=f"OASIS+ {self._format_domain_name(domain_type)} Domain",
            unrestricted_threshold=thresholds.get("unrestricted", 42),
            small_business_threshold=thresholds.get("small_business", 36),
            criteria=criteria_list,
        )

        return domain

    def _find_headers(self, sheet: Worksheet) -> Tuple[int, Dict[str, int]]:
        """Find the header row and map column positions"""
        col_map = {}

        # Search first 10 rows for header
        for row_idx in range(1, min(11, sheet.max_row + 1)):
            row_values = []
            for col_idx in range(1, min(20, sheet.max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = str(cell.value).lower() if cell.value else ""
                row_values.append((col_idx, value))

            # Check if this looks like a header row
            potential_map = {}
            for col_idx, value in row_values:
                if not value:
                    continue

                # Check for criteria ID column
                if any(re.search(p, value) for p in CRITERIA_ID_PATTERNS):
                    potential_map["criteria_id"] = col_idx

                # Check for description column
                if any(re.search(p, value) for p in DESCRIPTION_PATTERNS):
                    potential_map["description"] = col_idx

                # Check for points column
                if any(re.search(p, value) for p in POINTS_PATTERNS):
                    potential_map["points"] = col_idx

                # Check for validation/rule column
                if "validation" in value or "rule" in value or "verify" in value:
                    potential_map["validation"] = col_idx

                # Check for type column
                if "type" in value or "category" in value:
                    potential_map["type"] = col_idx

            # If we found at least criteria_id or description, use this row
            if "criteria_id" in potential_map or "description" in potential_map:
                col_map = potential_map
                return row_idx, col_map

        return 0, {}

    def _parse_criteria_row(
        self,
        sheet: Worksheet,
        row_idx: int,
        col_map: Dict[str, int],
        domain_type: DomainType
    ) -> Optional[ScoringCriteria]:
        """Parse a single criteria row"""

        def get_cell_value(col_name: str) -> Any:
            if col_name in col_map:
                return sheet.cell(row=row_idx, column=col_map[col_name]).value
            return None

        criteria_id = get_cell_value("criteria_id")
        description = get_cell_value("description")
        points = get_cell_value("points")
        validation = get_cell_value("validation")
        criteria_type_str = get_cell_value("type")

        # Skip empty rows
        if not criteria_id and not description:
            return None

        # Clean up values
        criteria_id = str(criteria_id).strip() if criteria_id else f"AUTO-{row_idx}"
        description = str(description).strip() if description else ""

        # Skip header-like rows or section titles
        if not description or len(description) < 10:
            return None

        # Parse points
        try:
            max_points = int(points) if points else 0
        except (ValueError, TypeError):
            max_points = 0

        # Determine criteria type
        criteria_type = self._determine_criteria_type(
            criteria_id, description, criteria_type_str
        )

        # Extract threshold value if mentioned
        threshold_value = self._extract_threshold_value(description)

        # Check if J.P-3 required
        requires_jp3 = bool(
            validation and ("j.p-3" in str(validation).lower() or "jp3" in str(validation).lower())
        ) or "verification form" in description.lower()

        return ScoringCriteria(
            criteria_id=criteria_id,
            domain=domain_type,
            description=description,
            max_points=max_points,
            criteria_type=criteria_type,
            validation_rule=str(validation) if validation else None,
            threshold_value=threshold_value,
            requires_jp3=requires_jp3,
        )

    def _determine_criteria_type(
        self,
        criteria_id: str,
        description: str,
        type_str: Optional[str]
    ) -> CriteriaType:
        """Determine the criteria type from available information"""

        desc_lower = description.lower()
        id_lower = criteria_id.lower()

        # Check explicit type string
        if type_str:
            type_lower = str(type_str).lower()
            if "mandatory" in type_lower:
                return CriteriaType.MANDATORY
            if "optional" in type_lower or "credit" in type_lower:
                return CriteriaType.OPTIONAL_CREDIT
            if "qualifying" in type_lower:
                return CriteriaType.QUALIFYING_PROJECT
            if "federal" in type_lower or "fep" in type_lower:
                return CriteriaType.FEDERAL_EXPERIENCE

        # Infer from description
        if "must" in desc_lower or "mandatory" in desc_lower or "required" in desc_lower:
            return CriteriaType.MANDATORY

        if "qualifying project" in desc_lower or "qp" in desc_lower:
            return CriteriaType.QUALIFYING_PROJECT

        if "federal experience" in desc_lower or "fep" in desc_lower:
            return CriteriaType.FEDERAL_EXPERIENCE

        if "threshold" in desc_lower or "minimum" in desc_lower:
            return CriteriaType.THRESHOLD

        # Default to optional credit
        return CriteriaType.OPTIONAL_CREDIT

    def _extract_threshold_value(self, description: str) -> Optional[Decimal]:
        """Extract dollar threshold values from description"""
        # Match patterns like $500,000 or $500K or 500000
        patterns = [
            r'\$\s*([\d,]+)\s*(?:k|K)',  # $500K
            r'\$\s*([\d,]+(?:\.\d{2})?)',  # $500,000.00
            r'([\d,]+)\s*(?:dollars?)',  # 500000 dollars
        ]

        for pattern in patterns:
            match = re.search(pattern, description)
            if match:
                value_str = match.group(1).replace(",", "")
                try:
                    value = Decimal(value_str)
                    # If K notation, multiply by 1000
                    if "k" in description[match.start():match.end()].lower():
                        value *= 1000
                    return value
                except Exception:
                    continue

        return None

    def _extract_thresholds(self, sheet: Worksheet) -> Dict[str, int]:
        """Extract threshold requirements from sheet"""
        thresholds = {
            "unrestricted": 42,
            "small_business": 36,
        }

        # Search for threshold mentions
        for row_idx in range(1, min(50, sheet.max_row + 1)):
            for col_idx in range(1, min(10, sheet.max_column + 1)):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if not cell.value:
                    continue

                value = str(cell.value).lower()
                if "threshold" in value or "minimum" in value:
                    # Look for number in adjacent cells
                    for adj_col in range(col_idx, min(col_idx + 3, sheet.max_column + 1)):
                        adj_cell = sheet.cell(row=row_idx, column=adj_col)
                        try:
                            points = int(adj_cell.value)
                            if 30 <= points <= 50:  # Reasonable threshold range
                                if "unrestricted" in value:
                                    thresholds["unrestricted"] = points
                                elif "small" in value:
                                    thresholds["small_business"] = points
                                else:
                                    # Default to unrestricted
                                    thresholds["unrestricted"] = points
                                break
                        except (ValueError, TypeError):
                            continue

        return thresholds

    def _format_domain_name(self, domain_type: DomainType) -> str:
        """Format domain type enum to readable name"""
        return domain_type.value.replace("_", " ").title()

    def get_domain(self, domain_type: DomainType) -> Optional[OASISDomain]:
        """Get a specific domain by type"""
        return self.domains.get(domain_type)

    def get_all_criteria(self) -> List[ScoringCriteria]:
        """Get all criteria across all domains"""
        all_criteria = []
        for domain in self.domains.values():
            all_criteria.extend(domain.criteria)
        return all_criteria

    def get_mandatory_criteria(self, domain_type: DomainType) -> List[ScoringCriteria]:
        """Get mandatory criteria for a domain"""
        domain = self.domains.get(domain_type)
        if not domain:
            return []
        return [c for c in domain.criteria if c.criteria_type == CriteriaType.MANDATORY]

    def get_optional_credits(self, domain_type: DomainType) -> List[ScoringCriteria]:
        """Get optional credit criteria for a domain"""
        domain = self.domains.get(domain_type)
        if not domain:
            return []
        return [c for c in domain.criteria if c.criteria_type == CriteriaType.OPTIONAL_CREDIT]

    def calculate_max_possible_score(self, domain_type: DomainType) -> int:
        """Calculate maximum possible score for a domain"""
        domain = self.domains.get(domain_type)
        if not domain:
            return 0
        return sum(c.max_points for c in domain.criteria)

    def export_to_dict(self) -> Dict[str, Any]:
        """Export parsed data as dictionary for serialization"""
        return {
            domain_type.value: {
                "name": domain.name,
                "description": domain.description,
                "thresholds": {
                    "unrestricted": domain.unrestricted_threshold,
                    "small_business": domain.small_business_threshold,
                },
                "criteria": [
                    {
                        "id": c.criteria_id,
                        "description": c.description,
                        "max_points": c.max_points,
                        "type": c.criteria_type.value,
                        "requires_jp3": c.requires_jp3,
                    }
                    for c in domain.criteria
                ]
            }
            for domain_type, domain in self.domains.items()
        }


def parse_jp1_matrix(filepath: str) -> Dict[DomainType, OASISDomain]:
    """
    Convenience function to parse a J.P-1 file.

    Args:
        filepath: Path to the J.P-1 Excel file

    Returns:
        Dictionary mapping DomainType to OASISDomain objects
    """
    parser = JP1MatrixParser()
    return parser.parse_file(filepath)
