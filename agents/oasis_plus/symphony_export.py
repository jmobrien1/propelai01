"""
Symphony Export Bundle Generator
================================

Generates complete submission bundles ready for upload to the
GSA Symphony (OSP) portal.

Bundle Contents:
- Tagged evidence PDFs with annotations
- J.P-3 verification forms
- Evidence index document
- Score summary report
- Checklist for manual review
"""

import logging
import os
import shutil
import json
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional, Any
from zipfile import ZipFile

try:
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.units import inch
    from reportlab.pdfgen import canvas
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, ListFlowable, ListItem
    )
    from reportlab.lib import colors
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .models import (
    OASISDomain,
    Project,
    ProjectClaim,
    ScorecardResult,
    DomainType,
    BusinessSize,
    VerificationStatus,
)
from .pdf_tagger import TaggedPDF
from .form_generator import JP3GenerationResult

logger = logging.getLogger(__name__)


@dataclass
class BundleManifest:
    """Manifest of bundle contents"""
    proposal_id: str
    contractor_name: str
    domain: str
    business_size: str
    generated_at: datetime

    # Score summary
    total_score: int
    threshold: int
    margin: int
    meets_threshold: bool

    # File counts
    evidence_pdfs: int = 0
    jp3_forms: int = 0

    # File list
    files: List[Dict[str, str]] = field(default_factory=list)


@dataclass
class BundleResult:
    """Result of bundle generation"""
    bundle_path: str
    manifest: BundleManifest
    success: bool = True
    errors: List[str] = field(default_factory=list)


class SymphonyBundleGenerator:
    """
    Generates complete Symphony submission bundles.

    Creates a ZIP archive containing all required files for
    OASIS+ proposal submission via the Symphony portal.
    """

    def __init__(self):
        if not REPORTLAB_AVAILABLE:
            logger.warning("ReportLab not available - some features disabled")
        if not OPENPYXL_AVAILABLE:
            logger.warning("openpyxl not available - Excel export disabled")

    def generate_bundle(
        self,
        output_path: str,
        proposal_id: str,
        contractor_name: str,
        domain: DomainType,
        business_size: BusinessSize,
        scorecard: ScorecardResult,
        tagged_pdfs: List[TaggedPDF],
        jp3_forms: List[JP3GenerationResult],
        domain_obj: Optional[OASISDomain] = None,
    ) -> BundleResult:
        """
        Generate a complete Symphony submission bundle.

        Args:
            output_path: Path for the output ZIP file
            proposal_id: Proposal identifier
            contractor_name: Name of contractor
            domain: Target domain
            business_size: Business size category
            scorecard: Scorecard with optimization results
            tagged_pdfs: List of tagged evidence PDFs
            jp3_forms: List of J.P-3 forms
            domain_obj: Optional domain object with criteria

        Returns:
            BundleResult with generation status
        """
        manifest = BundleManifest(
            proposal_id=proposal_id,
            contractor_name=contractor_name,
            domain=domain.value,
            business_size=business_size.value,
            generated_at=datetime.now(),
            total_score=scorecard.total_score,
            threshold=scorecard.threshold,
            margin=scorecard.margin,
            meets_threshold=scorecard.meets_threshold,
        )

        result = BundleResult(
            bundle_path=output_path,
            manifest=manifest,
        )

        # Create temp directory for bundle contents
        bundle_dir = Path(output_path).with_suffix('')
        bundle_dir.mkdir(parents=True, exist_ok=True)

        try:
            # 1. Copy tagged PDFs
            evidence_dir = bundle_dir / "01_Evidence_PDFs"
            evidence_dir.mkdir(exist_ok=True)

            for pdf in tagged_pdfs:
                if Path(pdf.output_path).exists():
                    dest = evidence_dir / Path(pdf.output_path).name
                    shutil.copy2(pdf.output_path, dest)
                    manifest.files.append({
                        "type": "evidence_pdf",
                        "path": f"01_Evidence_PDFs/{dest.name}",
                        "annotations": pdf.annotations_added,
                    })
                    manifest.evidence_pdfs += 1

            # 2. Copy J.P-3 forms
            forms_dir = bundle_dir / "02_JP3_Forms"
            forms_dir.mkdir(exist_ok=True)

            for form in jp3_forms:
                if form.success and Path(form.output_path).exists():
                    dest = forms_dir / Path(form.output_path).name
                    shutil.copy2(form.output_path, dest)
                    manifest.files.append({
                        "type": "jp3_form",
                        "path": f"02_JP3_Forms/{dest.name}",
                        "project": form.project_title,
                    })
                    manifest.jp3_forms += 1

            # 3. Generate evidence index
            if REPORTLAB_AVAILABLE:
                index_path = bundle_dir / "03_Evidence_Index.pdf"
                self._generate_evidence_index(
                    str(index_path),
                    scorecard,
                    tagged_pdfs,
                    contractor_name,
                    domain,
                )
                manifest.files.append({
                    "type": "index",
                    "path": "03_Evidence_Index.pdf",
                })

            # 4. Generate score summary report
            if REPORTLAB_AVAILABLE:
                summary_path = bundle_dir / "04_Score_Summary.pdf"
                self._generate_score_summary(
                    str(summary_path),
                    scorecard,
                    contractor_name,
                    domain,
                    business_size,
                    domain_obj,
                )
                manifest.files.append({
                    "type": "summary",
                    "path": "04_Score_Summary.pdf",
                })

            # 5. Generate Excel checklist
            if OPENPYXL_AVAILABLE:
                checklist_path = bundle_dir / "05_Submission_Checklist.xlsx"
                self._generate_checklist(
                    str(checklist_path),
                    scorecard,
                    tagged_pdfs,
                    jp3_forms,
                    domain,
                )
                manifest.files.append({
                    "type": "checklist",
                    "path": "05_Submission_Checklist.xlsx",
                })

            # 6. Generate manifest JSON
            manifest_path = bundle_dir / "MANIFEST.json"
            with open(manifest_path, 'w') as f:
                json.dump({
                    "proposal_id": manifest.proposal_id,
                    "contractor_name": manifest.contractor_name,
                    "domain": manifest.domain,
                    "business_size": manifest.business_size,
                    "generated_at": manifest.generated_at.isoformat(),
                    "score": {
                        "total": manifest.total_score,
                        "threshold": manifest.threshold,
                        "margin": manifest.margin,
                        "meets_threshold": manifest.meets_threshold,
                    },
                    "contents": {
                        "evidence_pdfs": manifest.evidence_pdfs,
                        "jp3_forms": manifest.jp3_forms,
                    },
                    "files": manifest.files,
                }, f, indent=2)

            # 7. Create ZIP archive
            with ZipFile(output_path, 'w') as zipf:
                for root, dirs, files in os.walk(bundle_dir):
                    for file in files:
                        file_path = Path(root) / file
                        arcname = file_path.relative_to(bundle_dir)
                        zipf.write(file_path, arcname)

            # Cleanup temp directory
            shutil.rmtree(bundle_dir)

            logger.info(f"Generated Symphony bundle: {output_path}")

        except Exception as e:
            result.success = False
            result.errors.append(str(e))
            logger.error(f"Bundle generation failed: {e}")

        return result

    def _generate_evidence_index(
        self,
        output_path: str,
        scorecard: ScorecardResult,
        tagged_pdfs: List[TaggedPDF],
        contractor_name: str,
        domain: DomainType,
    ):
        """Generate PDF evidence index document"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=16,
            alignment=1,
            spaceAfter=20,
        )
        story.append(Paragraph(
            f"OASIS+ Evidence Index<br/>{domain.value.replace('_', ' ').title()} Domain",
            title_style
        ))

        story.append(Paragraph(f"<b>Contractor:</b> {contractor_name}", styles['Normal']))
        story.append(Paragraph(
            f"<b>Generated:</b> {datetime.now().strftime('%Y-%m-%d %H:%M')}",
            styles['Normal']
        ))
        story.append(Spacer(1, 20))

        # Score summary
        story.append(Paragraph("<b>Score Summary</b>", styles['Heading2']))
        score_data = [
            ["Total Score:", str(scorecard.total_score)],
            ["Threshold:", str(scorecard.threshold)],
            ["Margin:", f"{scorecard.margin:+d}"],
            ["Status:", "QUALIFIES" if scorecard.meets_threshold else "BELOW THRESHOLD"],
        ]
        score_table = Table(score_data, colWidths=[2*inch, 2*inch])
        score_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ]))
        story.append(score_table)
        story.append(Spacer(1, 20))

        # Evidence index by document
        story.append(Paragraph("<b>Tagged Evidence Documents</b>", styles['Heading2']))
        story.append(Spacer(1, 10))

        for pdf in tagged_pdfs:
            story.append(Paragraph(
                f"<b>{Path(pdf.output_path).name}</b>",
                styles['Heading3']
            ))
            story.append(Paragraph(
                f"Annotations: {pdf.annotations_added} | "
                f"Claims Tagged: {len(pdf.claims_tagged)} | "
                f"Pages Modified: {len(pdf.pages_modified)}",
                styles['Normal']
            ))

            if pdf.claims_tagged:
                claim_list = ListFlowable(
                    [ListItem(Paragraph(c, styles['Normal'])) for c in pdf.claims_tagged[:10]],
                    bulletType='bullet',
                )
                story.append(claim_list)

            story.append(Spacer(1, 10))

        doc.build(story)

    def _generate_score_summary(
        self,
        output_path: str,
        scorecard: ScorecardResult,
        contractor_name: str,
        domain: DomainType,
        business_size: BusinessSize,
        domain_obj: Optional[OASISDomain],
    ):
        """Generate PDF score summary report"""
        doc = SimpleDocTemplate(
            output_path,
            pagesize=letter,
            rightMargin=0.75*inch,
            leftMargin=0.75*inch,
            topMargin=0.75*inch,
            bottomMargin=0.75*inch,
        )

        styles = getSampleStyleSheet()
        story = []

        # Title
        title_style = ParagraphStyle(
            'Title',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=1,
            spaceAfter=20,
        )
        story.append(Paragraph("OASIS+ Score Summary Report", title_style))

        # Header info
        header_data = [
            ["Contractor:", contractor_name],
            ["Domain:", domain.value.replace('_', ' ').title()],
            ["Business Size:", business_size.value.replace('_', ' ').title()],
            ["Report Date:", datetime.now().strftime('%Y-%m-%d')],
        ]
        header_table = Table(header_data, colWidths=[1.5*inch, 4*inch])
        header_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 20))

        # Score box
        story.append(Paragraph("<b>QUALIFICATION STATUS</b>", styles['Heading2']))

        status_color = colors.green if scorecard.meets_threshold else colors.red
        status_text = "QUALIFIES" if scorecard.meets_threshold else "DOES NOT QUALIFY"

        status_data = [
            [f"Total Score: {scorecard.total_score}", f"Threshold: {scorecard.threshold}"],
            [f"Margin: {scorecard.margin:+d}", f"Status: {status_text}"],
        ]
        status_table = Table(status_data, colWidths=[3*inch, 3*inch])
        status_table.setStyle(TableStyle([
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('BOX', (0, 0), (-1, -1), 2, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        story.append(status_table)
        story.append(Spacer(1, 20))

        # Points breakdown
        story.append(Paragraph("<b>Points Breakdown</b>", styles['Heading2']))
        breakdown_data = [
            ["Category", "Points"],
            ["Verified Points", str(scorecard.verified_points)],
            ["Pending Verification", str(scorecard.pending_points)],
            ["Unverified", str(scorecard.unverified_points)],
            ["Total", str(scorecard.total_score)],
        ]
        breakdown_table = Table(breakdown_data, colWidths=[3*inch, 1.5*inch])
        breakdown_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
            ('ALIGN', (1, 0), (1, -1), 'CENTER'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(breakdown_table)
        story.append(Spacer(1, 20))

        # Selected projects
        story.append(Paragraph("<b>Selected Qualifying Projects</b>", styles['Heading2']))

        project_data = [["#", "Project Title", "Agency", "AAV"]]
        for idx, project in enumerate(scorecard.qualifying_projects, 1):
            aav = project.calculate_aav()
            project_data.append([
                str(idx),
                project.title[:40] + "..." if len(project.title) > 40 else project.title,
                project.client_agency,
                f"${aav:,.0f}",
            ])

        project_table = Table(project_data, colWidths=[0.5*inch, 3*inch, 1.5*inch, 1*inch])
        project_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),
            ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
            ('BOX', (0, 0), (-1, -1), 1, colors.black),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        story.append(project_table)
        story.append(Spacer(1, 20))

        # Risk factors
        if scorecard.at_risk_claims:
            story.append(Paragraph("<b>Risk Factors</b>", styles['Heading2']))
            story.append(Paragraph(
                f"There are {len(scorecard.at_risk_claims)} claims with low confidence "
                "that should be reviewed before submission.",
                styles['Normal']
            ))

        doc.build(story)

    def _generate_checklist(
        self,
        output_path: str,
        scorecard: ScorecardResult,
        tagged_pdfs: List[TaggedPDF],
        jp3_forms: List[JP3GenerationResult],
        domain: DomainType,
    ):
        """Generate Excel submission checklist"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Submission Checklist"

        # Styles
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, size=12, color="FFFFFF")

        # Title
        ws['A1'] = f"OASIS+ Submission Checklist - {domain.value.replace('_', ' ').title()}"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:E1')

        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A3'] = f"Score: {scorecard.total_score} / Threshold: {scorecard.threshold}"

        # Checklist items
        row = 5
        ws[f'A{row}'] = "Item"
        ws[f'B{row}'] = "Description"
        ws[f'C{row}'] = "Status"
        ws[f'D{row}'] = "File"
        ws[f'E{row}'] = "Notes"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}{row}'].font = header_font_white
            ws[f'{col}{row}'].fill = header_fill

        row += 1

        # Pre-submission checks
        checks = [
            ("1", "J.P-1 Matrix Reviewed", "☐", "", "Verify all criteria claimed"),
            ("2", "Projects Meet AAV", "☐", "", "Minimum $500K (Unrestricted) or $250K (SB)"),
            ("3", "Recency Verified", "☐", "", "Projects within 5 years"),
            ("4", "Evidence PDFs Tagged", "☐", f"{len(tagged_pdfs)} files", "All tags visible in Symphony"),
            ("5", "J.P-3 Forms Signed", "☐", f"{len(jp3_forms)} forms", "CO signatures obtained"),
            ("6", "Score Above Threshold", "☑" if scorecard.meets_threshold else "☐",
             f"{scorecard.total_score} pts", f"Need {scorecard.threshold} minimum"),
        ]

        for item in checks:
            ws[f'A{row}'] = item[0]
            ws[f'B{row}'] = item[1]
            ws[f'C{row}'] = item[2]
            ws[f'D{row}'] = item[3]
            ws[f'E{row}'] = item[4]
            row += 1

        row += 1

        # Evidence files
        ws[f'A{row}'] = "Evidence Files"
        ws[f'A{row}'].font = header_font
        row += 1

        for idx, pdf in enumerate(tagged_pdfs, 1):
            ws[f'A{row}'] = f"PDF-{idx}"
            ws[f'B{row}'] = Path(pdf.output_path).name
            ws[f'C{row}'] = f"{pdf.annotations_added} tags"
            ws[f'D{row}'] = "☐ Uploaded"
            row += 1

        row += 1

        # J.P-3 forms
        ws[f'A{row}'] = "J.P-3 Forms"
        ws[f'A{row}'].font = header_font
        row += 1

        for idx, form in enumerate(jp3_forms, 1):
            ws[f'A{row}'] = f"JP3-{idx}"
            ws[f'B{row}'] = form.project_title
            ws[f'C{row}'] = "☐ Signed" if form.success else "⚠ Error"
            ws[f'D{row}'] = "☐ Uploaded"
            row += 1

        # Auto-width columns
        for col_idx, width in enumerate([8, 40, 15, 20, 35], 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(output_path)


def generate_symphony_bundle(
    output_path: str,
    proposal_id: str,
    contractor_name: str,
    domain: DomainType,
    business_size: BusinessSize,
    scorecard: ScorecardResult,
    tagged_pdfs: List[TaggedPDF],
    jp3_forms: List[JP3GenerationResult],
    domain_obj: Optional[OASISDomain] = None,
) -> BundleResult:
    """
    Convenience function to generate a Symphony bundle.

    Returns:
        BundleResult with generation status
    """
    generator = SymphonyBundleGenerator()
    return generator.generate_bundle(
        output_path=output_path,
        proposal_id=proposal_id,
        contractor_name=contractor_name,
        domain=domain,
        business_size=business_size,
        scorecard=scorecard,
        tagged_pdfs=tagged_pdfs,
        jp3_forms=jp3_forms,
        domain_obj=domain_obj,
    )
